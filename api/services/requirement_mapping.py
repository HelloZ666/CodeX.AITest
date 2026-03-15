from __future__ import annotations

import io
import uuid
from collections import OrderedDict
from typing import Iterable

from services.file_parser import detect_file_type

TEMPLATE_HEADERS = ("标签", "需求关键字", "关联场景")
SOURCE_TYPES = {"upload", "manual", "mixed"}
OLE2_SIGNATURE = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"

TEMPLATE_GROUPS = [
    {
        "tag": "流程变更",
        "requirement_keyword": "抄录",
        "related_scenarios": ["一键抄录", "逐字抄录", "逐字点击", "文本框输入"],
    },
    {
        "tag": "模板变更",
        "requirement_keyword": "单证",
        "related_scenarios": ["在途订单投保", "ca单证核对", "本人、非本人签字模式", "重签重录"],
    },
    {
        "tag": "页面新增",
        "requirement_keyword": "新增页面",
        "related_scenarios": ["兼容性测试", "跳入页面的其他跳转页面相关性", "在途订单投保"],
    },
    {
        "tag": "弹窗",
        "requirement_keyword": "新增弹窗",
        "related_scenarios": ["弹窗内容核对", "弹窗页面其他弹窗相关性测试"],
    },
]


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip()
    return text


def _unique_preserve_order(values: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        normalized_value = _normalize_text(value)
        if not normalized_value or normalized_value in seen:
            continue
        seen.add(normalized_value)
        result.append(normalized_value)
    return result


def _resolve_excel_subtype(filename: str, content: bytes) -> str:
    lowered = filename.lower()
    if lowered.endswith(".xls"):
        return "xls"
    if lowered.endswith(".xlsx"):
        return "xlsx"
    if content.startswith(OLE2_SIGNATURE):
        return "xls"
    return "xlsx"


def _iter_xlsx_sheet_rows(worksheet) -> list[list[str]]:
    merged_lookup: dict[tuple[int, int], str] = {}
    for merged_range in worksheet.merged_cells.ranges:
        top_left_value = _normalize_text(worksheet.cell(merged_range.min_row, merged_range.min_col).value)
        for row_idx in range(merged_range.min_row, merged_range.max_row + 1):
            for col_idx in range(merged_range.min_col, merged_range.max_col + 1):
                merged_lookup[(row_idx, col_idx)] = top_left_value

    rows: list[list[str]] = []
    for row_idx in range(1, worksheet.max_row + 1):
        values = []
        for col_idx in range(1, worksheet.max_column + 1):
            raw_value = worksheet.cell(row_idx, col_idx).value
            if raw_value is None and (row_idx, col_idx) in merged_lookup:
                raw_value = merged_lookup[(row_idx, col_idx)]
            values.append(_normalize_text(raw_value))
        if any(values):
            rows.append(values)
    return rows


def _iter_xls_sheet_rows(sheet) -> list[list[str]]:
    merged_lookup: dict[tuple[int, int], str] = {}
    for row_low, row_high, col_low, col_high in getattr(sheet, "merged_cells", []):
        top_left_value = _normalize_text(sheet.cell_value(row_low, col_low))
        for row_idx in range(row_low, row_high):
            for col_idx in range(col_low, col_high):
                merged_lookup[(row_idx, col_idx)] = top_left_value

    rows: list[list[str]] = []
    for row_idx in range(sheet.nrows):
        values = []
        for col_idx in range(sheet.ncols):
            raw_value = sheet.cell_value(row_idx, col_idx)
            normalized_value = _normalize_text(raw_value)
            if not normalized_value and (row_idx, col_idx) in merged_lookup:
                normalized_value = merged_lookup[(row_idx, col_idx)]
            values.append(normalized_value)
        if any(values):
            rows.append(values)
    return rows


def _load_first_non_empty_sheet(filename: str, content: bytes) -> tuple[str, str, list[list[str]]]:
    excel_subtype = _resolve_excel_subtype(filename, content)
    if excel_subtype == "xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ImportError("openpyxl 库未安装，无法解析 xlsx 文件") from exc

        workbook = load_workbook(filename=io.BytesIO(content), read_only=False, data_only=True)
        try:
            for worksheet in workbook.worksheets:
                rows = _iter_xlsx_sheet_rows(worksheet)
                if rows:
                    return excel_subtype, worksheet.title, rows
        finally:
            workbook.close()
    else:
        try:
            import xlrd
        except ImportError as exc:
            raise ImportError("xlrd 库未安装，无法解析 xls 文件") from exc

        workbook = xlrd.open_workbook(file_contents=content, formatting_info=True)
        for sheet in workbook.sheets():
            rows = _iter_xls_sheet_rows(sheet)
            if rows:
                return excel_subtype, sheet.name, rows

    raise ValueError("Excel 文件为空")


def normalize_requirement_mapping_groups(groups: Iterable[dict]) -> list[dict]:
    if not isinstance(groups, list):
        raise ValueError("需求映射数据格式无效")

    grouped: OrderedDict[tuple[str, str], dict] = OrderedDict()
    for raw_group in groups:
        if not isinstance(raw_group, dict):
            raise ValueError("需求映射分组格式无效")

        tag = _normalize_text(raw_group.get("tag"))
        requirement_keyword = _normalize_text(raw_group.get("requirement_keyword"))
        related_scenarios_raw = raw_group.get("related_scenarios")

        if not isinstance(related_scenarios_raw, list):
            raise ValueError("关联场景必须为数组")

        related_scenarios = _unique_preserve_order(
            [
                normalized_value
                for item in related_scenarios_raw
                if (normalized_value := _normalize_text(item))
            ]
        )

        if not tag:
            raise ValueError("标签不能为空")
        if not requirement_keyword:
            raise ValueError("需求关键字不能为空")
        if not related_scenarios:
            raise ValueError("每个需求映射分组至少需要一条关联场景")

        group_key = (tag, requirement_keyword)
        existing = grouped.get(group_key)
        if existing is None:
            grouped[group_key] = {
                "id": _normalize_text(raw_group.get("id")) or uuid.uuid4().hex,
                "tag": tag,
                "requirement_keyword": requirement_keyword,
                "related_scenarios": list(related_scenarios),
            }
        else:
            existing["related_scenarios"] = _unique_preserve_order(
                [*existing["related_scenarios"], *related_scenarios]
            )

    return list(grouped.values())


def flatten_requirement_mapping_groups(groups: list[dict]) -> list[dict]:
    rows: list[dict] = []
    for group in groups:
        scenarios = group.get("related_scenarios") or []
        row_span = len(scenarios)
        for index, scenario in enumerate(scenarios):
            rows.append(
                {
                    "group_id": group["id"],
                    "row_key": f"{group['id']}-{index}",
                    "tag": group["tag"],
                    "requirement_keyword": group["requirement_keyword"],
                    "related_scenario": scenario,
                    "tag_row_span": row_span if index == 0 else 0,
                    "requirement_keyword_row_span": row_span if index == 0 else 0,
                    "operation_row_span": row_span if index == 0 else 0,
                }
            )
    return rows


def parse_requirement_mapping_file(filename: str, content: bytes) -> dict:
    if detect_file_type(filename) != "excel":
        raise ValueError("仅支持 Excel 文件")

    excel_subtype, sheet_name, rows = _load_first_non_empty_sheet(filename, content)
    header_row = rows[0]
    header = header_row[:3] + [""] * max(0, 3 - len(header_row[:3]))

    if tuple(header[:3]) != TEMPLATE_HEADERS:
        raise ValueError("模板表头必须为：标签、需求关键字、关联场景")
    if any(_normalize_text(value) for value in header_row[3:]):
        raise ValueError("模板仅支持三列：标签、需求关键字、关联场景")

    grouped_rows: list[dict] = []
    current_tag = ""
    current_keyword = ""

    for row_index, raw_row in enumerate(rows[1:], start=2):
        if any(_normalize_text(value) for value in raw_row[3:]):
            raise ValueError(f"第 {row_index} 行存在多余列数据，模板仅支持三列")

        row = raw_row[:3] + [""] * max(0, 3 - len(raw_row[:3]))
        tag, requirement_keyword, related_scenario = row[:3]

        if not any((tag, requirement_keyword, related_scenario)):
            continue

        if tag or requirement_keyword:
            if not tag or not requirement_keyword:
                raise ValueError(f"第 {row_index} 行缺少标签或需求关键字")
            current_tag = tag
            current_keyword = requirement_keyword
        elif not current_tag or not current_keyword:
            raise ValueError(f"第 {row_index} 行缺少标签或需求关键字")

        if not related_scenario:
            raise ValueError(f"第 {row_index} 行缺少关联场景")

        grouped_rows.append(
            {
                "tag": current_tag,
                "requirement_keyword": current_keyword,
                "related_scenarios": [related_scenario],
            }
        )

    groups = normalize_requirement_mapping_groups(grouped_rows)
    if not groups:
        raise ValueError("需求映射文件没有可导入的数据")

    flattened_rows = flatten_requirement_mapping_groups(groups)
    return {
        "excel_subtype": excel_subtype,
        "sheet_name": sheet_name,
        "groups": groups,
        "rows": flattened_rows,
        "group_count": len(groups),
        "row_count": len(flattened_rows),
    }


def build_requirement_mapping_template() -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError as exc:
        raise ImportError("openpyxl 库未安装，无法生成需求映射模板") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(list(TEMPLATE_HEADERS))

    for column_index, header in enumerate(TEMPLATE_HEADERS, start=1):
        cell = sheet.cell(row=1, column=column_index, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    current_row = 2
    for group in TEMPLATE_GROUPS:
        scenarios = group["related_scenarios"]
        start_row = current_row
        for scenario in scenarios:
            sheet.cell(row=current_row, column=3, value=scenario)
            current_row += 1

        end_row = current_row - 1
        sheet.cell(row=start_row, column=1, value=group["tag"])
        sheet.cell(row=start_row, column=2, value=group["requirement_keyword"])

        if end_row > start_row:
            sheet.merge_cells(start_row=start_row, end_row=end_row, start_column=1, end_column=1)
            sheet.merge_cells(start_row=start_row, end_row=end_row, start_column=2, end_column=2)

        for row_index in range(start_row, end_row + 1):
            for column_index in range(1, 4):
                sheet.cell(row=row_index, column=column_index).alignment = Alignment(
                    horizontal="center" if column_index < 3 else "left",
                    vertical="center",
                )

    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 18
    sheet.column_dimensions["C"].width = 36

    workbook.create_sheet(title="Sheet2")
    workbook.create_sheet(title="Sheet3")

    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()
