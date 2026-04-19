"""
file_parser.py - 文件解析工具模块

支持 CSV、Excel(xlsx)、JSON 和 Word(DOC/DOCX) 格式文件的基础识别。
"""

import csv
import io
import json
import zipfile
from typing import Iterable, Union

OLE2_SIGNATURE = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
WORD_FILE_TYPES = {"doc", "docx"}


def parse_csv(content: Union[str, bytes]) -> list[dict]:
    """
    解析 CSV 文件内容。

    Args:
        content: CSV 文件内容（字符串或字节）

    Returns:
        解析后的字典列表

    Raises:
        ValueError: CSV 格式无效
    """
    if isinstance(content, bytes):
        for encoding in ["utf-8", "gbk", "gb2312", "utf-8-sig"]:
            try:
                content = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise ValueError("无法识别CSV文件编码，请使用UTF-8编码")

    if not content.strip():
        raise ValueError("CSV文件内容为空")

    reader = csv.DictReader(io.StringIO(content))
    rows = list(reader)

    if not rows:
        raise ValueError("CSV文件没有数据行")

    return rows


def _is_empty_excel_row(row: Iterable[object]) -> bool:
    return not any(value is not None and str(value).strip() for value in row)


def _normalize_excel_headers(row: Iterable[object]) -> list[str]:
    headers = []
    for index, value in enumerate(row):
        header = str(value).strip() if value is not None else ""
        headers.append(header or f"col_{index}")
    return headers


def _parse_worksheet_rows(worksheet) -> tuple[list[str] | None, list[dict]]:
    headers = None
    parsed_rows: list[dict] = []

    for row in worksheet.iter_rows(values_only=True):
        if _is_empty_excel_row(row):
            continue

        if headers is None:
            headers = _normalize_excel_headers(row)
            continue

        row_dict = {}
        for index, value in enumerate(row):
            if index < len(headers):
                row_dict[headers[index]] = str(value).strip() if value is not None else ""

        if any(row_dict.values()):
            parsed_rows.append(row_dict)

    return headers, parsed_rows


def parse_excel(content: bytes) -> list[dict]:
    """
    解析 Excel(xlsx) 文件内容。

    Args:
        content: Excel 文件字节内容

    Returns:
        解析后的字典列表

    Raises:
        ValueError: Excel 格式无效
        ImportError: openpyxl 未安装
    """
    if content.startswith(OLE2_SIGNATURE):
        try:
            import xlrd
        except ImportError as exc:
            raise ImportError("xlrd库未安装，无法解析 xls 文件") from exc

        try:
            workbook = xlrd.open_workbook(file_contents=content)
        except Exception as exc:
            raise ValueError(f"Excel文件格式无效: {exc}") from exc

        found_headers = False
        for sheet in workbook.sheets():
            headers = None
            rows: list[dict] = []

            for row_index in range(sheet.nrows):
                row = [sheet.cell_value(row_index, col_index) for col_index in range(sheet.ncols)]
                if _is_empty_excel_row(row):
                    continue

                if headers is None:
                    headers = _normalize_excel_headers(row)
                    found_headers = True
                    continue

                row_dict = {}
                for index, value in enumerate(row):
                    if index < len(headers):
                        row_dict[headers[index]] = str(value).strip() if value is not None else ""

                if any(row_dict.values()):
                    rows.append(row_dict)

            if rows:
                return rows

        if found_headers:
            raise ValueError("Excel文件没有数据行")
        raise ValueError("Excel文件为空")

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError("openpyxl库未安装，无法解析Excel文件") from exc

    try:
        # 部分导出的 Excel 会带错误的 worksheet dimension。
        # read_only 模式会直接信任该元数据，导致整张表被截断为 A1。
        workbook = load_workbook(filename=io.BytesIO(content), read_only=False, data_only=True)
    except Exception as exc:
        raise ValueError(f"Excel文件格式无效: {exc}") from exc

    try:
        if not workbook.worksheets:
            raise ValueError("Excel文件为空")

        found_headers = False
        for worksheet in workbook.worksheets:
            headers, rows = _parse_worksheet_rows(worksheet)
            if headers:
                found_headers = True
            if rows:
                return rows

        if found_headers:
            raise ValueError("Excel文件没有数据行")
        raise ValueError("Excel文件为空")
    finally:
        workbook.close()


def parse_json(content: Union[str, bytes]) -> dict:
    """
    解析 JSON 文件内容。

    Args:
        content: JSON 文件内容（字符串或字节）

    Returns:
        解析后的字典

    Raises:
        ValueError: JSON 格式无效
    """
    if isinstance(content, bytes):
        try:
            content = content.decode("utf-8")
        except UnicodeDecodeError:
            content = content.decode("utf-8-sig")

    if not content.strip():
        raise ValueError("JSON文件内容为空")

    try:
        data = json.loads(content)
    except json.JSONDecodeError as exc:
        raise ValueError(f"JSON格式无效: {exc}") from exc

    return data


def detect_file_type(filename: str) -> str:
    """
    根据文件名检测文件类型。

    Args:
        filename: 文件名

    Returns:
        文件类型: "csv", "excel", "json", "doc", "docx", "pdf", "yaml", "unknown"
    """
    name_lower = filename.lower()
    if name_lower.endswith(".csv"):
        return "csv"
    if name_lower.endswith((".xlsx", ".xls")):
        return "excel"
    if name_lower.endswith(".json"):
        return "json"
    if name_lower.endswith(".pdf"):
        return "pdf"
    if name_lower.endswith((".md", ".markdown")):
        return "markdown"
    if name_lower.endswith((".yaml", ".yml")):
        return "yaml"
    if name_lower.endswith(".doc"):
        return "doc"
    if name_lower.endswith(".docx"):
        return "docx"
    return "unknown"


def detect_word_content_type(content: bytes) -> str | None:
    if not content:
        return None

    if content.startswith(OLE2_SIGNATURE):
        return "doc"

    if zipfile.is_zipfile(io.BytesIO(content)):
        return "docx"

    return None


def _validate_word_content(content: bytes) -> str:
    if not content:
        return "Word文档内容为空"

    word_type = detect_word_content_type(content)
    if word_type is None:
        return "当前文件不是有效的 Word 文档，请确认上传的是标准 .docx 或旧版 .doc 文档"

    return ""


def validate_file(filename: str, content: bytes, allowed_types: list[str], max_size_mb: float = 10.0) -> str:
    """
    校验上传文件的类型和大小。

    Args:
        filename: 文件名
        content: 文件内容字节
        allowed_types: 允许的文件类型列表，如 ["csv", "excel", "json", "doc", "docx"]
        max_size_mb: 最大文件大小（MB）

    Returns:
        错误信息，为空字符串表示校验通过
    """
    file_type = detect_file_type(filename)
    word_type_allowed = bool(WORD_FILE_TYPES.intersection(allowed_types))
    actual_word_type = detect_word_content_type(content) if word_type_allowed else None

    if file_type == "unknown":
        if actual_word_type and actual_word_type in allowed_types:
            file_type = actual_word_type
        else:
            return f"不支持的文件格式: {filename}，请上传 {', '.join(allowed_types)} 格式文件"

    if file_type in WORD_FILE_TYPES and actual_word_type in WORD_FILE_TYPES:
        file_type = actual_word_type

    if file_type == "unknown":
        return f"不支持的文件格式: {filename}，请上传 {', '.join(allowed_types)} 格式文件"

    if file_type not in allowed_types:
        return f"该接口不支持 {file_type} 格式，请上传 {', '.join(allowed_types)} 格式文件"

    size_mb = len(content) / (1024 * 1024)
    if size_mb > max_size_mb:
        return f"文件过大 ({size_mb:.1f}MB)，最大允许 {max_size_mb}MB"

    if file_type in WORD_FILE_TYPES:
        word_err = _validate_word_content(content)
        if word_err:
            return word_err

    return ""
