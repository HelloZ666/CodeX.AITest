"""Workbook parsing and aggregation for the efficiency analysis dashboard."""

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import date, datetime
from typing import Iterable

from services.file_parser import OLE2_SIGNATURE

BUSINESS_LIFE = "寿险"
BUSINESS_HEALTH = "健康险"
SUPPORTED_BUSINESSES = (BUSINESS_LIFE, BUSINESS_HEALTH)

HISTORY_SHEET_NAMES: dict[str, str] = {
    BUSINESS_LIFE: "寿险汇总数据-历年数据",
    BUSINESS_HEALTH: "健康险汇总数据-历年数据",
}

CURRENT_SHEET_PATTERNS: tuple[tuple[re.Pattern[str], str, str, int], ...] = (
    (re.compile(r"^寿险汇总数据-含外协（(?P<year>20\d{2})）$"), BUSINESS_LIFE, "summary", 30),
    (re.compile(r"^寿险汇总-含外协（(?P<year>20\d{2})）$"), BUSINESS_LIFE, "summary", 20),
    (re.compile(r"^健康险汇总数据-含外协（(?P<year>20\d{2})）$"), BUSINESS_HEALTH, "summary", 30),
    (re.compile(r"^健康险数据-含外协（(?P<year>20\d{2})）$"), BUSINESS_HEALTH, "summary", 20),
    (re.compile(r"^寿险对外数据（(?P<year>20\d{2})）$"), BUSINESS_LIFE, "external", 30),
    (re.compile(r"^健康险对外数据（(?P<year>20\d{2})）$"), BUSINESS_HEALTH, "external", 30),
    (re.compile(r"^各团队数据（寿险）-?(?P<year>20\d{2})\s*$"), BUSINESS_LIFE, "team", 30),
    (re.compile(r"^各团队数据（健康险）-?(?P<year>20\d{2})\s*$"), BUSINESS_HEALTH, "team", 30),
)

MONTH_RE = re.compile(r"(?P<month>1[0-2]|[1-9])(?:月|月份)")
YEAR_MONTH_VALUE_RE = re.compile(
    r"(?P<year>20\d{2})\s*(?:年|[-/.])\s*(?P<month>1[0-2]|0?[1-9])(?:\s*月|\s*(?:[-/.]\s*\d{1,2})?)",
)
YEAR_LABEL_RE = re.compile(r"(?P<year>20\d{2})年")
SHORT_YEAR_RE = re.compile(r"(?P<year>\d{2})年")
TEAM_BLOCK_RE = re.compile(r"(?P<year>20\d{2})年(?P<month>1[0-2]|[1-9])月份")
HEADER_CLEAN_RE = re.compile(r"[\s\n\r\t()（）【】\[\]{}:：、,，/\\+_\-]")


@dataclass
class WorkbookSheet:
    name: str
    rows: list[list[object]]


SUMMARY_ALIASES = {
    "sync_tasks": ["同步任务数"],
    "total_tasks": ["同步+回归", "同步任务数回归任务数", "同步回归"],
    "release_count": ["发布总次数"],
    "demand_count": ["需求数", "需求数(同步+需求号去重)"],
    "defect_count": ["缺陷数", "缺陷数同步回归+安全", "缺陷数 除性能、代码扫描外所有任务类型（SIT+FT）"],
    "avg_cycle_days": ["测试任务平均时效"],
    "design_cases": ["设计用例数", "设计案例数"],
    "execution_cases": ["执行案例数", "执行用例数"],
    "functional_manpower": ["功能人月投入", "人月投入"],
    "performance_manpower": ["性能人月投入"],
    "qa_manpower": ["QA人月投入"],
}

EXTERNAL_ALIASES = {
    "manpower_input": ["人月投入"],
    "release_count": ["发布总次数"],
    "sync_tasks": ["同步任务数"],
    "demand_count": ["需求数"],
    "design_cases": ["设计案例数", "设计用例数"],
    "execution_cases": ["执行案例数", "执行用例数"],
    "defect_rate": ["缺陷率"],
    "avg_cycle_days": ["测试任务平均时效"],
    "defect_count": ["测试缺陷数", "缺陷数"],
    "production_defect_count": ["生产缺陷数"],
    "production_defect_detection_rate": ["生产缺陷检出率"],
    "automation_coverage": ["自动化覆盖率"],
    "automation_pass_rate": ["自动化执行通过率"],
    "planned_app_count": ["计划应用数"],
    "connected_app_count": ["已接入应用数"],
    "precision_access_rate": ["精准接入率"],
}

TEAM_ALIASES = {
    "system_count": ["系统个数"],
    "sync_tasks": ["同步任务数"],
    "total_tasks": ["同步任务数回归任务数", "同步+回归"],
    "demand_count": ["总需求数", "需求数"],
    "bug_count": ["BUG数", "BUG数同步任务+回归+安全"],
    "total_bug_count": ["总BUG数", "总BUG数全任务"],
    "design_cases": ["设计案例数", "设计案例数同步任务"],
    "execution_cases": ["执行案例数", "执行案例数同步+回归+安全", "执行案例数同步+回归(手工&自动化)+安全"],
    "staff_count": ["人数", "人月数"],
    "per_capita_task": ["人均任务", "人均任务_数量"],
    "per_capita_task_rank": ["人均任务_排名"],
    "per_capita_demand": ["人均需求数", "人均需求数_数量"],
    "per_capita_demand_rank": ["人均需求数_排名"],
    "per_capita_bug": ["人均缺陷数", "人均缺陷数_数量"],
    "per_capita_bug_rank": ["人均缺陷数_排名"],
    "defect_rate": ["缺陷率", "缺陷率_数量"],
    "defect_rate_rank": ["缺陷率_排名"],
    "avg_design_cases": ["平均设计案例数", "平均设计案例数_数量"],
    "avg_design_cases_rank": ["平均设计案例数_排名"],
    "avg_execution_cases": ["平均执行案例数", "平均执行案例数_数量"],
    "avg_execution_cases_rank": ["平均执行案例数_排名"],
}

PER_CAPITA_ALIASES = {
    "per_capita_sync_tasks": ["人均同步任务"],
    "per_capita_total_tasks": ["人均同步+回归", "人均(同步+回归)"],
    "per_capita_demand_count": ["人均需求数"],
    "per_capita_defect_count": ["人均总缺陷数", "人均缺陷数"],
    "defect_rate": ["缺陷率"],
    "avg_design_cases": ["平均设计案例数"],
    "avg_execution_cases": ["执行案例数", "执行案例数同步+回归+安全"],
}


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\u3000", " ").strip()


def _normalize_header(value: object) -> str:
    return HEADER_CLEAN_RE.sub("", _clean_text(value))


def _is_empty_row(row: Iterable[object]) -> bool:
    return not any(_clean_text(v) for v in row)


def _coerce_number(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    text = _clean_text(value)
    if not text or text in {"无数据", "-", "--", "/"}:
        return None
    text = text.replace(",", "").replace("，", "")
    if text.endswith("%"):
        try:
            return float(text[:-1]) / 100
        except ValueError:
            return None
    try:
        return float(text)
    except ValueError:
        return None


def _coerce_int(value: object) -> int | None:
    number = _coerce_number(value)
    return None if number is None else int(round(number))


def _round(value: float | None, digits: int = 4) -> float | None:
    return None if value is None else round(value, digits)


def _extract_month(value: object) -> int | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.month
    if isinstance(value, date):
        return value.month
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        month = int(value)
        return month if 1 <= month <= 12 else None
    text = _clean_text(value)
    match = MONTH_RE.search(text)
    if match:
        return int(match.group("month"))
    year_month_match = YEAR_MONTH_VALUE_RE.search(text)
    if year_month_match:
        return int(year_month_match.group("month"))
    if text.isdigit():
        month = int(text)
        return month if 1 <= month <= 12 else None
    return None


def _extract_year(value: object) -> int | None:
    text = _clean_text(value)
    full = YEAR_LABEL_RE.search(text)
    if full:
        return int(full.group("year"))
    short = SHORT_YEAR_RE.search(text)
    if short:
        return 2000 + int(short.group("year"))
    return None


def _header_index(rows: list[list[object]], start: int = 0) -> int | None:
    for idx in range(start, len(rows)):
        if any(_normalize_header(v) == "月份" for v in rows[idx]):
            return idx
    return None


def _header_map(row: list[object]) -> dict[str, int]:
    result: dict[str, int] = {}
    for idx, value in enumerate(row):
        normalized = _normalize_header(value)
        if normalized and normalized not in result:
            result[normalized] = idx
    return result


def _pick(row: list[object], hmap: dict[str, int], aliases: list[str]) -> object:
    for alias in aliases:
        key = _normalize_header(alias)
        col = hmap.get(key)
        if col is not None and col < len(row):
            return row[col]
    return None


def _has_metrics(record: dict[str, object], fields: list[str]) -> bool:
    return any(record.get(f) is not None for f in fields)


def _merge_non_null_fields(
    target: dict[str, object],
    incoming: dict[str, object],
    *,
    skip_keys: set[str] | None = None,
) -> None:
    skip = skip_keys or set()
    for key, value in incoming.items():
        if key in skip or value is None:
            continue
        if target.get(key) is None:
            target[key] = value


def _merge_current_records_by_month(
    candidates: list[tuple[int, list[list[object]]]],
    parser,
    business: str,
    year: int,
) -> list[dict[str, object]]:
    merged: dict[int, dict[str, object]] = {}
    for _, rows in sorted(candidates, key=lambda item: item[0], reverse=True):
        for record in parser(rows, business, year):
            month = int(record["month"])
            if month not in merged:
                merged[month] = dict(record)
                continue
            _merge_non_null_fields(
                merged[month],
                record,
                skip_keys={"business", "year", "month", "month_label"},
            )
    return [merged[month] for month in sorted(merged)]


def _merge_team_rows_by_name(
    target_rows: list[dict[str, object]],
    incoming_rows: list[dict[str, object]],
) -> list[dict[str, object]]:
    merged = {str(row["team_name"]): dict(row) for row in target_rows}
    for row in incoming_rows:
        team_name = str(row["team_name"])
        if team_name not in merged:
            merged[team_name] = dict(row)
            continue
        _merge_non_null_fields(merged[team_name], row, skip_keys={"team_name"})
    return list(merged.values())


def _merge_current_team_snapshots(
    candidates: list[tuple[int, list[list[object]]]],
    business: str,
    year: int,
) -> list[dict[str, object]]:
    merged: dict[int, dict[str, object]] = {}
    for _, rows in sorted(candidates, key=lambda item: item[0], reverse=True):
        for snapshot in _parse_current_team(rows, business):
            if int(snapshot["year"]) != year:
                continue
            month = int(snapshot["month"])
            teams = [dict(team) for team in snapshot["teams"]]
            if month not in merged:
                merged[month] = {**snapshot, "teams": teams}
                continue
            target = merged[month]
            _merge_non_null_fields(
                target,
                {key: value for key, value in snapshot.items() if key != "teams"},
                skip_keys={"business", "year", "month", "month_label"},
            )
            target["teams"] = _merge_team_rows_by_name(target["teams"], teams)
    return [merged[month] for month in sorted(merged)]


def _row_width(row: list[object]) -> int:
    for idx in range(len(row) - 1, -1, -1):
        if _clean_text(row[idx]):
            return idx + 1
    return 0


def _slice_row(row: list[object], width: int) -> list[object]:
    values = list(row[:width])
    if len(values) < width:
        values.extend([None] * (width - len(values)))
    return values


def _build_raw_history_table(
    rows: list[list[object]],
    header_idx: int,
    start_idx: int,
    end_idx: int | None,
) -> dict[str, object]:
    header_row = rows[header_idx] if 0 <= header_idx < len(rows) else []
    width = _row_width(header_row)
    if width <= 0:
        return {"headers": [], "rows": []}

    limit = len(rows) if end_idx is None else min(end_idx, len(rows))
    headers = [_clean_text(value) for value in _slice_row(header_row, width)]
    table_rows: list[list[object]] = []
    for row in rows[start_idx:limit]:
        if _extract_year(row[0] if row else None) is None:
            continue
        table_rows.append(_slice_row(row, width))

    keep_indices = [
        idx
        for idx, header in enumerate(headers)
        if _clean_text(header) or any(_clean_text(row[idx] if idx < len(row) else None) for row in table_rows)
    ]
    if keep_indices and len(keep_indices) != len(headers):
        headers = [headers[idx] for idx in keep_indices]
        table_rows = [
            [row[idx] if idx < len(row) else None for idx in keep_indices]
            for row in table_rows
        ]

    return {"headers": headers, "rows": table_rows}


def _classify(name: str) -> dict[str, object] | None:
    clean = name.strip()
    for business, sheet in HISTORY_SHEET_NAMES.items():
        if clean == sheet:
            return {"scope": "history", "business": business, "kind": "summary", "year": None, "score": 100}
    for pattern, business, kind, score in CURRENT_SHEET_PATTERNS:
        match = pattern.fullmatch(clean)
        if match:
            return {"scope": "current", "business": business, "kind": kind, "year": int(match.group("year")), "score": score}
    return None


def _parse_current_summary(rows: list[list[object]], business: str, year: int) -> list[dict[str, object]]:
    idx = _header_index(rows)
    if idx is None:
        return []
    hmap = _header_map(rows[idx])
    records: list[dict[str, object]] = []
    for row in rows[idx + 1:]:
        first = _clean_text(row[0] if row else None)
        if "月度人均任务数据" in first:
            break
        month = _extract_month(row[0] if row else None)
        if month is None:
            continue
        record = {
            "business": business,
            "year": year,
            "month": month,
            "month_label": f"{month}月",
            "sync_tasks": _coerce_int(_pick(row, hmap, SUMMARY_ALIASES["sync_tasks"])),
            "total_tasks": _coerce_int(_pick(row, hmap, SUMMARY_ALIASES["total_tasks"])),
            "release_count": _coerce_int(_pick(row, hmap, SUMMARY_ALIASES["release_count"])),
            "demand_count": _coerce_int(_pick(row, hmap, SUMMARY_ALIASES["demand_count"])),
            "defect_count": _coerce_int(_pick(row, hmap, SUMMARY_ALIASES["defect_count"])),
            "avg_cycle_days": _round(_coerce_number(_pick(row, hmap, SUMMARY_ALIASES["avg_cycle_days"])), 2),
            "design_cases": _coerce_int(_pick(row, hmap, SUMMARY_ALIASES["design_cases"])),
            "execution_cases": _coerce_int(_pick(row, hmap, SUMMARY_ALIASES["execution_cases"])),
            "functional_manpower": _round(_coerce_number(_pick(row, hmap, SUMMARY_ALIASES["functional_manpower"])), 2),
            "performance_manpower": _round(_coerce_number(_pick(row, hmap, SUMMARY_ALIASES["performance_manpower"])), 2),
            "qa_manpower": _round(_coerce_number(_pick(row, hmap, SUMMARY_ALIASES["qa_manpower"])), 2),
        }
        if _has_metrics(record, ["sync_tasks", "total_tasks", "demand_count", "defect_count"]):
            records.append(record)
    return records


def _parse_current_external(rows: list[list[object]], business: str, year: int) -> list[dict[str, object]]:
    idx = _header_index(rows)
    if idx is None:
        return []
    hmap = _header_map(rows[idx])
    records: list[dict[str, object]] = []
    for row in rows[idx + 1:]:
        month = _extract_month(row[0] if row else None)
        if month is None:
            continue
        record = {
            "business": business,
            "year": year,
            "month": month,
            "month_label": f"{month}月",
            "manpower_input": _round(_coerce_number(_pick(row, hmap, EXTERNAL_ALIASES["manpower_input"])), 2),
            "release_count": _coerce_int(_pick(row, hmap, EXTERNAL_ALIASES["release_count"])),
            "sync_tasks": _coerce_int(_pick(row, hmap, EXTERNAL_ALIASES["sync_tasks"])),
            "demand_count": _coerce_int(_pick(row, hmap, EXTERNAL_ALIASES["demand_count"])),
            "design_cases": _coerce_int(_pick(row, hmap, EXTERNAL_ALIASES["design_cases"])),
            "execution_cases": _coerce_int(_pick(row, hmap, EXTERNAL_ALIASES["execution_cases"])),
            "defect_rate": _round(_coerce_number(_pick(row, hmap, EXTERNAL_ALIASES["defect_rate"])), 6),
            "avg_cycle_days": _round(_coerce_number(_pick(row, hmap, EXTERNAL_ALIASES["avg_cycle_days"])), 2),
            "defect_count": _coerce_int(_pick(row, hmap, EXTERNAL_ALIASES["defect_count"])),
            "production_defect_count": _coerce_int(_pick(row, hmap, EXTERNAL_ALIASES["production_defect_count"])),
            "production_defect_detection_rate": _round(_coerce_number(_pick(row, hmap, EXTERNAL_ALIASES["production_defect_detection_rate"])), 6),
            "automation_coverage": _round(_coerce_number(_pick(row, hmap, EXTERNAL_ALIASES["automation_coverage"])), 6),
            "automation_pass_rate": _round(_coerce_number(_pick(row, hmap, EXTERNAL_ALIASES["automation_pass_rate"])), 6),
            "planned_app_count": _coerce_int(_pick(row, hmap, EXTERNAL_ALIASES["planned_app_count"])),
            "connected_app_count": _coerce_int(_pick(row, hmap, EXTERNAL_ALIASES["connected_app_count"])),
            "precision_access_rate": _round(_coerce_number(_pick(row, hmap, EXTERNAL_ALIASES["precision_access_rate"])), 6),
        }
        if _has_metrics(record, ["sync_tasks", "demand_count", "defect_rate", "defect_count"]):
            records.append(record)
    return records


def _team_headers(header_row: list[object], sub_row: list[object]) -> dict[str, int]:
    result: dict[str, int] = {}
    max_len = max(len(header_row), len(sub_row))
    for idx in range(max_len):
        main = _clean_text(header_row[idx]) if idx < len(header_row) else ""
        sub = _clean_text(sub_row[idx]) if idx < len(sub_row) else ""
        value = f"{main}_{sub}" if main and sub in {"数量", "排名"} else (main or sub or f"col_{idx}")
        key = _normalize_header(value)
        if key and key not in result:
            result[key] = idx
    return result


def _parse_current_team(rows: list[list[object]], business: str) -> list[dict[str, object]]:
    snapshots: list[dict[str, object]] = []
    idx = 0
    while idx < len(rows):
        first = _clean_text(rows[idx][0] if rows[idx] else None)
        match = TEAM_BLOCK_RE.search(first)
        if not match:
            idx += 1
            continue
        year, month = int(match.group("year")), int(match.group("month"))
        if idx + 2 >= len(rows):
            break
        hmap = _team_headers(rows[idx + 1], rows[idx + 2])
        teams: list[dict[str, object]] = []
        row_idx = idx + 3
        while row_idx < len(rows):
            row = rows[row_idx]
            team_name = _clean_text(row[0] if row else None)
            if not team_name or TEAM_BLOCK_RE.search(team_name):
                break
            record = {
                "team_name": team_name,
                "system_count": _coerce_int(_pick(row, hmap, TEAM_ALIASES["system_count"])),
                "sync_tasks": _coerce_int(_pick(row, hmap, TEAM_ALIASES["sync_tasks"])),
                "total_tasks": _coerce_int(_pick(row, hmap, TEAM_ALIASES["total_tasks"])),
                "demand_count": _coerce_int(_pick(row, hmap, TEAM_ALIASES["demand_count"])),
                "bug_count": _coerce_int(_pick(row, hmap, TEAM_ALIASES["bug_count"])),
                "total_bug_count": _coerce_int(_pick(row, hmap, TEAM_ALIASES["total_bug_count"])),
                "design_cases": _coerce_int(_pick(row, hmap, TEAM_ALIASES["design_cases"])),
                "execution_cases": _coerce_int(_pick(row, hmap, TEAM_ALIASES["execution_cases"])),
                "staff_count": _round(_coerce_number(_pick(row, hmap, TEAM_ALIASES["staff_count"])), 2),
                "per_capita_task": _round(_coerce_number(_pick(row, hmap, TEAM_ALIASES["per_capita_task"])), 4),
                "per_capita_task_rank": _coerce_int(_pick(row, hmap, TEAM_ALIASES["per_capita_task_rank"])),
                "per_capita_demand": _round(_coerce_number(_pick(row, hmap, TEAM_ALIASES["per_capita_demand"])), 4),
                "per_capita_demand_rank": _coerce_int(_pick(row, hmap, TEAM_ALIASES["per_capita_demand_rank"])),
                "per_capita_bug": _round(_coerce_number(_pick(row, hmap, TEAM_ALIASES["per_capita_bug"])), 4),
                "per_capita_bug_rank": _coerce_int(_pick(row, hmap, TEAM_ALIASES["per_capita_bug_rank"])),
                "defect_rate": _round(_coerce_number(_pick(row, hmap, TEAM_ALIASES["defect_rate"])), 6),
                "defect_rate_rank": _coerce_int(_pick(row, hmap, TEAM_ALIASES["defect_rate_rank"])),
                "avg_design_cases": _round(_coerce_number(_pick(row, hmap, TEAM_ALIASES["avg_design_cases"])), 2),
                "avg_design_cases_rank": _coerce_int(_pick(row, hmap, TEAM_ALIASES["avg_design_cases_rank"])),
                "avg_execution_cases": _round(_coerce_number(_pick(row, hmap, TEAM_ALIASES["avg_execution_cases"])), 2),
                "avg_execution_cases_rank": _coerce_int(_pick(row, hmap, TEAM_ALIASES["avg_execution_cases_rank"])),
            }
            if _has_metrics(record, ["sync_tasks", "demand_count", "total_bug_count", "staff_count"]):
                teams.append(record)
            row_idx += 1
        if teams:
            snapshots.append({"business": business, "year": year, "month": month, "month_label": f"{month}月", "teams": teams})
        next_row_first = _clean_text(rows[row_idx][0] if row_idx < len(rows) and rows[row_idx] else None)
        idx = row_idx if TEAM_BLOCK_RE.search(next_row_first) else row_idx + 1
    return snapshots


def _parse_history(rows: list[list[object]], business: str) -> dict[str, object]:
    idx = _header_index(rows)
    if idx is None:
        return {
            "business": business,
            "available_years": [],
            "latest_year": None,
            "yearly_summary": [],
            "yearly_per_capita": [],
            "yearly_summary_table": {"headers": [], "rows": []},
            "yearly_per_capita_table": {"headers": [], "rows": []},
        }
    hmap = _header_map(rows[idx])
    yearly_summary: list[dict[str, object]] = []
    split_idx: int | None = None
    for i in range(idx + 1, len(rows)):
        row = rows[i]
        first = _clean_text(row[0] if row else None)
        if "月度人均任务数据" in first:
            split_idx = i
            break
        year = _extract_year(first)
        if year is None:
            continue
        defect_count = _round(_coerce_number(_pick(row, hmap, SUMMARY_ALIASES["defect_count"])), 2)
        design_cases = _round(_coerce_number(_pick(row, hmap, SUMMARY_ALIASES["design_cases"])), 2)
        record = {
            "business": business,
            "year": year,
            "sync_tasks": _round(_coerce_number(_pick(row, hmap, SUMMARY_ALIASES["sync_tasks"])), 2),
            "total_tasks": _round(_coerce_number(_pick(row, hmap, SUMMARY_ALIASES["total_tasks"])), 2),
            "release_count": _round(_coerce_number(_pick(row, hmap, SUMMARY_ALIASES["release_count"])), 2),
            "demand_count": _round(_coerce_number(_pick(row, hmap, SUMMARY_ALIASES["demand_count"])), 2),
            "defect_count": defect_count,
            "defect_rate": _round(defect_count / design_cases, 6) if defect_count is not None and design_cases not in (None, 0) else None,
            "avg_cycle_days": _round(_coerce_number(_pick(row, hmap, SUMMARY_ALIASES["avg_cycle_days"])), 2),
            "design_cases": design_cases,
            "execution_cases": _round(_coerce_number(_pick(row, hmap, SUMMARY_ALIASES["execution_cases"])), 2),
            "functional_manpower": _round(_coerce_number(_pick(row, hmap, SUMMARY_ALIASES["functional_manpower"])), 2),
            "performance_manpower": _round(_coerce_number(_pick(row, hmap, SUMMARY_ALIASES["performance_manpower"])), 2),
            "qa_manpower": _round(_coerce_number(_pick(row, hmap, SUMMARY_ALIASES["qa_manpower"])), 2),
        }
        if _has_metrics(record, ["sync_tasks", "total_tasks", "demand_count", "defect_count"]):
            yearly_summary.append(record)

    summary_table = _build_raw_history_table(rows, idx, idx + 1, split_idx)
    per_capita_table: dict[str, object] = {"headers": [], "rows": []}
    yearly_per_capita: list[dict[str, object]] = []
    if split_idx is not None:
        p_idx = _header_index(rows, split_idx + 1)
        if p_idx is not None:
            per_capita_table = _build_raw_history_table(rows, p_idx, p_idx + 1, None)
            p_map = _header_map(rows[p_idx])
            for row in rows[p_idx + 1:]:
                year = _extract_year(row[0] if row else None)
                if year is None:
                    continue
                record = {
                    "business": business,
                    "year": year,
                    "per_capita_sync_tasks": _round(_coerce_number(_pick(row, p_map, PER_CAPITA_ALIASES["per_capita_sync_tasks"])), 4),
                    "per_capita_total_tasks": _round(_coerce_number(_pick(row, p_map, PER_CAPITA_ALIASES["per_capita_total_tasks"])), 4),
                    "per_capita_demand_count": _round(_coerce_number(_pick(row, p_map, PER_CAPITA_ALIASES["per_capita_demand_count"])), 4),
                    "per_capita_defect_count": _round(_coerce_number(_pick(row, p_map, PER_CAPITA_ALIASES["per_capita_defect_count"])), 4),
                    "defect_rate": _round(_coerce_number(_pick(row, p_map, PER_CAPITA_ALIASES["defect_rate"])), 6),
                    "avg_design_cases": _round(_coerce_number(_pick(row, p_map, PER_CAPITA_ALIASES["avg_design_cases"])), 4),
                    "avg_execution_cases": _round(_coerce_number(_pick(row, p_map, PER_CAPITA_ALIASES["avg_execution_cases"])), 4),
                }
                if _has_metrics(record, ["per_capita_sync_tasks", "per_capita_total_tasks", "defect_rate"]):
                    yearly_per_capita.append(record)
    yearly_summary.sort(key=lambda x: int(x["year"]))
    yearly_per_capita.sort(key=lambda x: int(x["year"]))
    years = sorted({int(x["year"]) for x in yearly_summary} | {int(x["year"]) for x in yearly_per_capita})
    return {
        "business": business,
        "available_years": years,
        "latest_year": years[-1] if years else None,
        "yearly_summary": yearly_summary,
        "yearly_per_capita": yearly_per_capita,
        "yearly_summary_table": summary_table,
        "yearly_per_capita_table": per_capita_table,
    }


def _merge_current(summary: list[dict[str, object]], external: list[dict[str, object]]) -> list[dict[str, object]]:
    merged = {int(item["month"]): dict(item) for item in summary}
    for ext in external:
        month = int(ext["month"])
        target = merged.setdefault(month, {"business": ext["business"], "year": ext["year"], "month": ext["month"], "month_label": ext["month_label"]})
        for k, v in ext.items():
            if k in {"business", "year", "month", "month_label"} or v is None:
                continue
            if target.get(k) is None:
                target[k] = v
    result: list[dict[str, object]] = []
    for month in sorted(merged):
        row = merged[month]
        if row.get("defect_rate") is None and row.get("defect_count") is not None and row.get("design_cases") not in (None, 0):
            row["defect_rate"] = _round(float(row["defect_count"]) / float(row["design_cases"]), 6)
        if row.get("functional_manpower") is None and row.get("manpower_input") is not None:
            row["functional_manpower"] = row["manpower_input"]
        result.append(row)
    return result


def _month_options(summary: list[dict[str, object]], external: list[dict[str, object]], teams: list[dict[str, object]]) -> list[dict[str, object]]:
    enabled = {int(item["month"]) for item in summary + external + teams}
    return [{"month": m, "month_label": f"{m}月", "has_data": m in enabled, "disabled": m not in enabled} for m in range(1, 13)]


def _months_payload(summary: list[dict[str, object]], external: list[dict[str, object]], teams: list[dict[str, object]]) -> dict[int, dict[str, object]]:
    s_map = {int(item["month"]): item for item in summary}
    e_map = {int(item["month"]): item for item in external}
    t_map = {int(item["month"]): item for item in teams}
    return {m: {"summary": s_map.get(m), "external": e_map.get(m), "team_snapshot": t_map.get(m)} for m in range(1, 13)}


def load_workbook_sheets(content: bytes) -> list[WorkbookSheet]:
    """Load all worksheet rows from an Excel workbook."""
    if content.startswith(OLE2_SIGNATURE):
        try:
            import xlrd
        except ImportError as exc:
            raise ImportError("xlrd 库未安装，无法解析 .xls 文件") from exc
        workbook = xlrd.open_workbook(file_contents=content)
        return [
            WorkbookSheet(
                name=sheet.name,
                rows=[[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)],
            )
            for sheet in workbook.sheets()
        ]
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError("openpyxl 库未安装，无法解析 Excel 文件") from exc
    workbook = load_workbook(filename=io.BytesIO(content), read_only=False, data_only=True)
    try:
        return [WorkbookSheet(name=ws.title, rows=[list(row) for row in ws.iter_rows(values_only=True)]) for ws in workbook.worksheets]
    finally:
        workbook.close()


def analyze_performance_workbook(content: bytes) -> dict[str, object]:
    """Parse efficiency-related workbook sheets into dashboard-friendly data."""
    sheets = load_workbook_sheets(content)
    if not sheets:
        raise ValueError("导入文件中未发现可解析的工作表")

    history_rows: dict[str, list[list[object]]] = {}
    current_candidates: dict[tuple[str, str, int], list[tuple[int, list[list[object]]]]] = {}
    classified: list[dict[str, object]] = []

    for sheet in sheets:
        if not sheet.rows or all(_is_empty_row(row) for row in sheet.rows):
            continue
        info = _classify(sheet.name)
        if info is None:
            continue
        classified.append(info)
        if info["scope"] == "history":
            history_rows[str(info["business"])] = sheet.rows
            continue
        key = (str(info["business"]), str(info["kind"]), int(info["year"]))
        score = int(info["score"])
        current_candidates.setdefault(key, []).append((score, sheet.rows))

    years = [int(item["year"]) for item in classified if item["scope"] == "current" and item["year"] is not None]
    current_year = max(years) if years else None

    businesses: dict[str, dict[str, object]] = {}
    for business in SUPPORTED_BUSINESSES:
        history = _parse_history(history_rows.get(business, []), business)
        summary: list[dict[str, object]] = []
        external: list[dict[str, object]] = []
        teams: list[dict[str, object]] = []
        if current_year is not None:
            summary_candidates = current_candidates.get((business, "summary", current_year), [])
            external_candidates = current_candidates.get((business, "external", current_year), [])
            team_candidates = current_candidates.get((business, "team", current_year), [])
            if summary_candidates:
                summary = _merge_current_records_by_month(
                    summary_candidates,
                    _parse_current_summary,
                    business,
                    current_year,
                )
            if external_candidates:
                external = _merge_current_records_by_month(
                    external_candidates,
                    _parse_current_external,
                    business,
                    current_year,
                )
            if team_candidates:
                teams = _merge_current_team_snapshots(team_candidates, business, current_year)
        summary.sort(key=lambda x: int(x["month"]))
        external.sort(key=lambda x: int(x["month"]))
        teams.sort(key=lambda x: int(x["month"]))
        month_options = _month_options(summary, external, teams)
        latest = max([m for m in month_options if m["has_data"]], key=lambda x: int(x["month"]), default=None)
        current = {
            "business": business,
            "year": current_year,
            "latest_month": (
                {"year": current_year, "month": int(latest["month"]), "month_label": str(latest["month_label"])}
                if latest and current_year is not None
                else None
            ),
            "month_options": month_options,
            "months": _months_payload(summary, external, teams),
        }
        if not history["available_years"] and not any(not m["disabled"] for m in month_options):
            continue
        businesses[business] = {"business": business, "history": history, "current": current}

    if not businesses:
        raise ValueError("导入文件中未识别出寿险或健康险效能分析数据")

    return {
        "available_businesses": list(businesses.keys()),
        "current_year": current_year,
        "businesses": businesses,
        "sheet_names": [sheet.name for sheet in sheets],
    }
