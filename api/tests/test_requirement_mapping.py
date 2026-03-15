import io

import pytest
from openpyxl import Workbook

from services.requirement_mapping import (
    build_requirement_mapping_template,
    normalize_requirement_mapping_groups,
    parse_requirement_mapping_file,
)


def build_requirement_mapping_xlsx_bytes(with_extra_column: bool = False) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    headers = ["标签", "需求关键字", "关联场景"]
    if with_extra_column:
        headers.append("额外字段")
    sheet.append(headers)
    sheet.merge_cells("A2:A3")
    sheet.merge_cells("B2:B3")
    sheet["A2"] = "流程变更"
    sheet["B2"] = "抄录"
    sheet["C2"] = "一键抄录"
    sheet["C3"] = "逐字抄录"

    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def build_requirement_mapping_xls_bytes() -> bytes:
    import xlwt

    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("Sheet1")
    sheet.write(0, 0, "标签")
    sheet.write(0, 1, "需求关键字")
    sheet.write(0, 2, "关联场景")
    sheet.write_merge(1, 2, 0, 0, "流程变更")
    sheet.write_merge(1, 2, 1, 1, "抄录")
    sheet.write(1, 2, "一键抄录")
    sheet.write(2, 2, "逐字抄录")

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


class TestParseRequirementMappingFile:
    def test_parse_xlsx_with_merged_cells(self):
        result = parse_requirement_mapping_file(
            "requirement-mapping.xlsx",
            build_requirement_mapping_xlsx_bytes(),
        )

        assert result["excel_subtype"] == "xlsx"
        assert result["sheet_name"] == "Sheet1"
        assert result["group_count"] == 1
        assert result["row_count"] == 2
        assert result["groups"][0]["tag"] == "流程变更"
        assert result["groups"][0]["requirement_keyword"] == "抄录"
        assert result["groups"][0]["related_scenarios"] == ["一键抄录", "逐字抄录"]

    def test_parse_xls_with_merged_cells(self):
        result = parse_requirement_mapping_file(
            "requirement-mapping.xls",
            build_requirement_mapping_xls_bytes(),
        )

        assert result["excel_subtype"] == "xls"
        assert result["group_count"] == 1
        assert result["groups"][0]["related_scenarios"] == ["一键抄录", "逐字抄录"]

    def test_parse_rejects_extra_columns(self):
        with pytest.raises(ValueError, match="仅支持三列"):
            parse_requirement_mapping_file(
                "requirement-mapping.xlsx",
                build_requirement_mapping_xlsx_bytes(with_extra_column=True),
            )

    def test_parse_rejects_invalid_header(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["标签", "关键字", "关联场景"])
        sheet.append(["流程变更", "抄录", "一键抄录"])

        buffer = io.BytesIO()
        workbook.save(buffer)
        workbook.close()

        with pytest.raises(ValueError, match="模板表头必须为"):
            parse_requirement_mapping_file("invalid.xlsx", buffer.getvalue())


class TestNormalizeRequirementMappingGroups:
    def test_normalize_merges_duplicate_groups(self):
        groups = normalize_requirement_mapping_groups(
            [
                {
                    "id": "group-1",
                    "tag": "流程变更",
                    "requirement_keyword": "抄录",
                    "related_scenarios": ["一键抄录", ""],
                },
                {
                    "id": "group-2",
                    "tag": "流程变更",
                    "requirement_keyword": "抄录",
                    "related_scenarios": ["逐字抄录"],
                },
            ]
        )

        assert groups == [
            {
                "id": "group-1",
                "tag": "流程变更",
                "requirement_keyword": "抄录",
                "related_scenarios": ["一键抄录", "逐字抄录"],
            }
        ]

    def test_normalize_rejects_missing_scenario(self):
        with pytest.raises(ValueError, match="至少需要一条关联场景"):
            normalize_requirement_mapping_groups(
                [
                    {
                        "tag": "流程变更",
                        "requirement_keyword": "抄录",
                        "related_scenarios": ["", " "],
                    }
                ]
            )


def test_build_requirement_mapping_template():
    content = build_requirement_mapping_template()
    result = parse_requirement_mapping_file("template.xlsx", content)

    assert result["group_count"] == 4
    assert result["row_count"] == 13
