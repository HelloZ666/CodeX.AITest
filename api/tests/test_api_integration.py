"""
test_api_integration.py - API集成测试（文件管理和分析记录路由）
"""

import io
import json
import os
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

from services.database import (
    init_db,
    save_analysis_record,
    save_requirement_analysis_record,
    upsert_external_user,
)


FIXTURES_DIR = Path(__file__).resolve().parent / "fixtures"


def build_code_mapping_xlsx_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "代码映射"
    sheet.append(["包名", "类名", "方法名", "功能描述"])
    sheet.append(["com.example.user", "UserService", "createUser", "创建新用户并发送欢迎邮件"])
    sheet.append(["com.example.user", "UserService", "updateUser", "更新用户基本信息"])

    content = io.BytesIO()
    workbook.save(content)
    workbook.close()
    return content.getvalue()


def build_real_test_cases_xlsx_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "用例列表"
    sheet.append([
        "",
        "流程名称说明",
        "功能模块路径说明",
        "用例描述说明",
        "预置条件说明",
        "测试步骤说明",
        "预期结果说明",
        "检查点类型说明",
        "测试类型说明",
        "用例等级说明",
        "用例类型说明",
        "用例优先级说明",
    ])
    sheet.append([
        "用例编号",
        "流程名称",
        "功能模块路径",
        "用例描述",
        "预置条件",
        "测试步骤",
        "预期结果",
        "检查点类型",
        "测试类型",
        "用例等级",
        "用例类型",
        "用例优先级",
    ])
    sheet.append([
        "case001",
        "用户流程_001",
        "用户中心-->用户管理-->删除用户",
        "删除用户正常流程验证",
        "系统已更新",
        "1、输入用户ID\n2、点击删除按钮\n3、检查返回结果",
        "删除用户成功，状态更新为deleted",
        "数据核对",
        "WEB功能测试",
        "核心",
        "正向",
        "P1",
    ])
    sheet.append([
        "case002",
        "订单流程_002",
        "订单中心-->订单管理-->订单总额",
        "计算订单总额验证",
        "系统已更新",
        "1、添加商品到订单\n2、提交订单\n3、检查总额",
        "订单总额计算正确",
        "数据核对",
        "WEB功能测试",
        "一般",
        "正向",
        "P2",
    ])

    content = io.BytesIO()
    workbook.save(content)
    workbook.close()
    return content.getvalue()


DEFECT_FIELDS = [
    "缺陷ID",
    "缺陷摘要",
    "任务编号",
    "系统名称",
    "系统CODE",
    "需求编号",
    "计划发布日期",
    "缺陷状态",
    "缺陷修复人",
    "缺陷修复人p13",
    "缺陷严重度",
    "重现频率",
    "业务影响",
    "缺陷来源",
    "缺陷原因",
    "缺陷子原因",
    "缺陷描述",
    "缺陷修复描述",
    "测试阶段",
    "分配处理人",
    "分配处理人P13",
    "缺陷修复时长",
    "修复轮次",
    "功能区",
    "缺陷关闭时间",
    "开发团队",
    "测试团队",
    "测试用例库",
    "功能模块",
    "测试项",
    "创建人姓名",
    "创建人P13",
    "创建时间",
    "是否初级缺陷",
    "初级缺陷依据",
]


@pytest.fixture(autouse=True)
def temp_db(tmp_path, monkeypatch):
    """为每个测试使用独立的临时数据库"""
    db_path = str(tmp_path / "test_api.db")
    monkeypatch.setattr("services.database.get_db_path", lambda: db_path)
    monkeypatch.setattr("services.production_issue_file_store.get_db_path", lambda: db_path)
    monkeypatch.setattr("services.test_issue_file_store.get_db_path", lambda: db_path)
    init_db()
    return db_path


@pytest.fixture
def client():
    """创建测试客户端"""
    from index import app
    with TestClient(app) as test_client:
        login_resp = test_client.post(
            "/api/auth/login",
            json={"username": "admin", "password": "Admin123!"},
        )
        assert login_resp.status_code == 200
        yield test_client


# ============ 项目CRUD ============

class TestProjectCRUD:
    """测试项目增删改查API"""

    def test_create_project(self, client):
        """创建项目"""
        resp = client.post("/api/projects", json={"name": "测试项目", "description": "描述"})
        assert resp.status_code == 200
        data = resp.json()
        assert data["success"] is True
        assert data["data"]["name"] == "测试项目"
        assert data["data"]["description"] == "描述"
        assert "id" in data["data"]

    def test_create_project_name_only(self, client):
        """仅提供名称创建项目"""
        resp = client.post("/api/projects", json={"name": "最小项目"})
        assert resp.status_code == 200
        assert resp.json()["data"]["description"] == ""

    def test_create_project_with_project_members(self, client):
        """创建项目时保存测试经理和测试人员"""
        manager = upsert_external_user(
            username="zhangyong-135",
            display_name="张勇",
            email="zhangyong-135@cpic.com.cn",
            external_profile={"deptname": "业务二部"},
        )
        tester = upsert_external_user(
            username="lisi-136",
            display_name="李四",
            email="lisi-136@cpic.com.cn",
            external_profile={"deptname": "业务一部"},
        )

        resp = client.post(
            "/api/projects",
            json={
                "name": "成员项目",
                "test_manager_ids": [manager["id"]],
                "tester_ids": [manager["id"], tester["id"]],
            },
        )

        assert resp.status_code == 200
        data = resp.json()["data"]
        assert data["test_manager_ids"] == [manager["id"]]
        assert data["tester_ids"] == [manager["id"], tester["id"]]

    def test_create_project_invalid_body(self, client):
        """缺少必填字段应返回422"""
        resp = client.post("/api/projects", json={"description": "没有名称"})
        assert resp.status_code == 422

    def test_list_projects_empty(self, client):
        """空数据库列出项目"""
        resp = client.get("/api/projects")
        assert resp.status_code == 200
        assert resp.json()["data"] == []

    def test_list_projects_with_data(self, client):
        """有数据时列出项目"""
        client.post("/api/projects", json={"name": "项目A"})
        client.post("/api/projects", json={"name": "项目B"})
        resp = client.get("/api/projects")
        assert resp.status_code == 200
        assert len(resp.json()["data"]) == 2

    def test_get_project(self, client):
        """获取项目详情（含统计信息）"""
        create_resp = client.post("/api/projects", json={"name": "项目"})
        project_id = create_resp.json()["data"]["id"]
        resp = client.get(f"/api/projects/{project_id}")
        assert resp.status_code == 200
        data = resp.json()["data"]
        assert data["name"] == "项目"
        assert "stats" in data
        assert data["stats"]["analysis_count"] == 0

    def test_get_project_not_found(self, client):
        """获取不存在的项目返回404"""
        resp = client.get("/api/projects/9999")
        assert resp.status_code == 404

    def test_update_project(self, client):
        """更新项目"""
        create_resp = client.post("/api/projects", json={"name": "原名称"})
        project_id = create_resp.json()["data"]["id"]
        manager = upsert_external_user(
            username="wangwu-137",
            display_name="王五",
            email="wangwu-137@cpic.com.cn",
            external_profile={"deptname": "业务三部"},
        )
        resp = client.put(
            f"/api/projects/{project_id}",
            json={"name": "新名称", "description": "新描述", "test_manager_ids": [manager["id"]]},
        )
        assert resp.status_code == 200
        data = resp.json()["data"]
        assert data["name"] == "新名称"
        assert data["description"] == "新描述"
        assert data["test_manager_ids"] == [manager["id"]]

    def test_create_project_rejects_non_p13_members(self, client):
        """测试经理和测试人员只能选择P13用户"""
        create_user_resp = client.post(
            "/api/users",
            json={
                "username": "local-user",
                "password": "Local12345!",
                "display_name": "本地用户",
                "role": "user",
            },
        )
        local_user_id = create_user_resp.json()["id"]

        resp = client.post(
            "/api/projects",
            json={
                "name": "非法成员项目",
                "test_manager_ids": [local_user_id],
            },
        )

        assert resp.status_code == 400

    def test_update_project_not_found(self, client):
        """更新不存在的项目返回404"""
        resp = client.put("/api/projects/9999", json={"name": "不存在"})
        assert resp.status_code == 404

    def test_delete_project(self, client):
        """删除项目"""
        create_resp = client.post("/api/projects", json={"name": "待删除"})
        project_id = create_resp.json()["data"]["id"]
        resp = client.delete(f"/api/projects/{project_id}")
        assert resp.status_code == 200
        assert resp.json()["success"] is True

        # 确认已删除
        get_resp = client.get(f"/api/projects/{project_id}")
        assert get_resp.status_code == 404

    def test_delete_project_not_found(self, client):
        """删除不存在的项目返回404"""
        resp = client.delete("/api/projects/9999")
        assert resp.status_code == 404


# ============ 映射文件上传 ============

class TestProjectMapping:
    """测试项目映射文件上传"""

    def test_upload_mapping(self, client):
        """上传映射文件到项目"""
        create_resp = client.post("/api/projects", json={"name": "项目"})
        project_id = create_resp.json()["data"]["id"]

        mapping_content = FIXTURES_DIR / "sample_mapping.csv"
        with open(mapping_content, "rb") as f:
            resp = client.post(
                f"/api/projects/{project_id}/mapping",
                files={"mapping_file": ("mapping.csv", f, "text/csv")},
            )
        assert resp.status_code == 200
        data = resp.json()
        assert data["success"] is True
        assert data["mapping_count"] > 0

        # 验证映射数据已保存
        get_resp = client.get(f"/api/projects/{project_id}")
        assert get_resp.json()["data"]["mapping_data"] is not None

    def test_upload_mapping_excel(self, client):
        """支持上传 Excel 代码映射文件"""
        create_resp = client.post("/api/projects", json={"name": "Excel项目"})
        project_id = create_resp.json()["data"]["id"]

        resp = client.post(
            f"/api/projects/{project_id}/mapping",
            files={
                "mapping_file": (
                    "mapping.xlsx",
                    build_code_mapping_xlsx_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert resp.status_code == 200
        data = resp.json()
        assert data["success"] is True
        assert data["mapping_count"] == 2
        assert data["data"]["mapping_data"][0]["package_name"] == "com.example.user"

    def test_create_project_mapping_entry(self, client):
        """支持手工追加单条项目代码映射"""
        create_resp = client.post("/api/projects", json={"name": "手工映射项目"})
        project_id = create_resp.json()["data"]["id"]

        resp = client.post(
            f"/api/projects/{project_id}/mapping/entries",
            json={
                "package_name": "com.example.order",
                "class_name": "OrderService",
                "method_name": "createOrder",
                "description": "创建订单并校验库存",
            },
        )

        assert resp.status_code == 200
        data = resp.json()
        assert data["success"] is True
        assert data["mapping_count"] == 1
        assert data["data"]["mapping_data"][0]["method_name"] == "createOrder"

    def test_create_project_mapping_entry_updates_duplicate(self, client):
        """重复的代码映射条目会覆盖更新原有描述"""
        create_resp = client.post("/api/projects", json={"name": "重复映射项目"})
        project_id = create_resp.json()["data"]["id"]

        first_resp = client.post(
            f"/api/projects/{project_id}/mapping/entries",
            json={
                "package_name": "com.example.order",
                "class_name": "OrderService",
                "method_name": "createOrder",
                "description": "创建订单并校验库存",
            },
        )
        assert first_resp.status_code == 200

        duplicate_resp = client.post(
            f"/api/projects/{project_id}/mapping/entries",
            json={
                "package_name": "com.example.order",
                "class_name": "OrderService",
                "method_name": "createOrder",
                "description": "重复描述",
            },
        )

        assert duplicate_resp.status_code == 200
        duplicate_data = duplicate_resp.json()
        assert duplicate_data["action"] == "updated"
        assert duplicate_data["mapping_count"] == 1
        assert duplicate_data["data"]["mapping_data"][0]["description"] == "重复描述"

    def test_update_project_mapping_entry(self, client):
        create_resp = client.post("/api/projects", json={"name": "编辑映射项目"})
        project_id = create_resp.json()["data"]["id"]

        client.post(
            f"/api/projects/{project_id}/mapping/entries",
            json={
                "package_name": "com.example.order",
                "class_name": "OrderService",
                "method_name": "createOrder",
                "description": "创建订单",
            },
        )

        update_resp = client.put(
            f"/api/projects/{project_id}/mapping/entries",
            json={
                "original_key": {
                    "package_name": "com.example.order",
                    "class_name": "OrderService",
                    "method_name": "createOrder",
                },
                "entry": {
                    "package_name": "com.example.order",
                    "class_name": "OrderService",
                    "method_name": "createOrder",
                    "description": "创建订单并校验库存",
                },
            },
        )

        assert update_resp.status_code == 200
        update_data = update_resp.json()
        assert update_data["action"] == "updated"
        assert update_data["mapping_count"] == 1
        assert update_data["data"]["mapping_data"][0]["description"] == "创建订单并校验库存"

    def test_update_project_mapping_entry_conflict(self, client):
        create_resp = client.post("/api/projects", json={"name": "冲突映射项目"})
        project_id = create_resp.json()["data"]["id"]

        client.post(
            f"/api/projects/{project_id}/mapping/entries",
            json={
                "package_name": "com.example.order",
                "class_name": "OrderService",
                "method_name": "createOrder",
                "description": "创建订单",
            },
        )
        client.post(
            f"/api/projects/{project_id}/mapping/entries",
            json={
                "package_name": "com.example.order",
                "class_name": "OrderService",
                "method_name": "cancelOrder",
                "description": "取消订单",
            },
        )

        conflict_resp = client.put(
            f"/api/projects/{project_id}/mapping/entries",
            json={
                "original_key": {
                    "package_name": "com.example.order",
                    "class_name": "OrderService",
                    "method_name": "createOrder",
                },
                "entry": {
                    "package_name": "com.example.order",
                    "class_name": "OrderService",
                    "method_name": "cancelOrder",
                    "description": "冲突写入",
                },
            },
        )

        assert conflict_resp.status_code == 409

    def test_delete_project_mapping_entry(self, client):
        create_resp = client.post("/api/projects", json={"name": "删除映射项目"})
        project_id = create_resp.json()["data"]["id"]

        client.post(
            f"/api/projects/{project_id}/mapping/entries",
            json={
                "package_name": "com.example.order",
                "class_name": "OrderService",
                "method_name": "createOrder",
                "description": "创建订单",
            },
        )

        delete_resp = client.delete(
            f"/api/projects/{project_id}/mapping/entries",
            params={
                "package_name": "com.example.order",
                "class_name": "OrderService",
                "method_name": "createOrder",
            },
        )

        assert delete_resp.status_code == 200
        delete_data = delete_resp.json()
        assert delete_data["action"] == "deleted"
        assert delete_data["mapping_count"] == 0

    def test_download_project_mapping_template(self, client):
        """支持下载 Excel 代码映射模板"""
        resp = client.get("/api/project-mapping-template")

        assert resp.status_code == 200
        assert resp.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert 'attachment; filename="code-mapping-template.xlsx"' in resp.headers["content-disposition"]

    def test_create_project_mapping_entry_with_test_point(self, client):
        create_resp = client.post("/api/projects", json={"name": "测试点映射项目"})
        project_id = create_resp.json()["data"]["id"]

        resp = client.post(
            f"/api/projects/{project_id}/mapping/entries",
            json={
                "package_name": "com.example.order",
                "class_name": "OrderService",
                "method_name": "createOrder",
                "description": "创建订单并校验库存",
                "test_point": "下单成功、库存不足、重复提交",
            },
        )

        assert resp.status_code == 200
        data = resp.json()["data"]["mapping_data"][0]
        assert data["test_point"] == "下单成功、库存不足、重复提交"

    def test_upload_mapping_excel_with_test_point(self, client):
        create_resp = client.post("/api/projects", json={"name": "测试点导入项目"})
        project_id = create_resp.json()["data"]["id"]

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "代码映射"
        sheet.append(["包名", "类名", "方法名", "功能描述", "测试点"])
        sheet.append([
            "com.example.order",
            "OrderService",
            "createOrder",
            "创建订单并校验库存",
            "下单成功、库存不足、重复提交",
        ])

        content = io.BytesIO()
        workbook.save(content)
        workbook.close()

        resp = client.post(
            f"/api/projects/{project_id}/mapping",
            files={
                "mapping_file": (
                    "mapping.xlsx",
                    content.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert resp.status_code == 200
        data = resp.json()["data"]["mapping_data"][0]
        assert data["test_point"] == "下单成功、库存不足、重复提交"

    def test_upload_mapping_project_not_found(self, client):
        """上传映射到不存在的项目返回404"""
        mapping_content = FIXTURES_DIR / "sample_mapping.csv"
        with open(mapping_content, "rb") as f:
            resp = client.post(
                "/api/projects/9999/mapping",
                files={"mapping_file": ("mapping.csv", f, "text/csv")},
            )
        assert resp.status_code == 404


# ============ 分析记录 ============

class TestRecords:
    """测试分析记录API"""

    def test_list_records_empty(self, client):
        """空数据库列出记录"""
        resp = client.get("/api/records")
        assert resp.status_code == 200
        assert resp.json()["data"] == []

    def test_get_record_not_found(self, client):
        """获取不存在的记录返回404"""
        resp = client.get("/api/records/9999")
        assert resp.status_code == 404

    def test_list_records_with_project_filter(self, client):
        """按项目过滤记录"""
        from services.database import create_project, save_analysis_record

        p1 = create_project(name="项目1")
        p2 = create_project(name="项目2")
        save_analysis_record(
            project_id=p1["id"],
            code_changes_summary={},
            test_coverage_result={},
            test_score=80.0,
            ai_suggestions=None,
            token_usage=0,
            cost=0.0,
            duration_ms=100,
        )
        save_analysis_record(
            project_id=p2["id"],
            code_changes_summary={},
            test_coverage_result={},
            test_score=90.0,
            ai_suggestions=None,
            token_usage=0,
            cost=0.0,
            duration_ms=200,
        )

        resp = client.get(f"/api/records?project_id={p1['id']}")
        assert resp.status_code == 200
        assert len(resp.json()["data"]) == 1


# ============ 项目分析 ============

class TestCaseQualityRecords:
    @staticmethod
    def _prepare_records(client):
        create_resp = client.post("/api/projects", json={"name": "案例质检项目"})
        project_id = create_resp.json()["data"]["id"]

        requirement_record = save_requirement_analysis_record(
            project_id=project_id,
            requirement_file_name="requirement.docx",
            section_snapshot={"selected_mode": "preferred_sections", "points": []},
            result_snapshot={
                "overview": {
                    "total_requirements": 6,
                    "matched_requirements": 3,
                    "mapping_hit_count": 3,
                    "unmatched_requirements": 3,
                },
                "score": {
                    "total_score": 72,
                    "grade": "C",
                    "summary": "需求评分示例",
                    "dimensions": [],
                },
                "mapping_suggestions": [
                    {
                        "requirement_point_id": "RP-1",
                        "section_number": "2.1",
                        "section_title": "下单流程",
                        "requirement_text": "提交订单",
                        "match_count": 1,
                        "suggestion": "补充库存不足、重复提交和订单回滚验证",
                    }
                ],
                "requirement_hits": [
                    {
                        "point_id": "RP-1",
                        "section_number": "2.1",
                        "section_title": "下单流程",
                        "text": "提交订单后需完成库存校验与订单落库。",
                        "mapping_suggestion": "补充库存不足、重复提交和订单回滚验证",
                        "mapping_matches": [
                            {
                                "tag": "下单主链路",
                                "requirement_keyword": "提交订单",
                                "matched_requirement_keyword": "提交订单",
                                "matched_scenarios": ["正常下单"],
                                "related_scenarios": ["库存不足", "重复提交"],
                                "additional_scenarios": ["订单回滚"],
                            }
                        ],
                    }
                ],
                "unmatched_requirements": [
                    {
                        "point_id": "RP-9",
                        "section_number": "4.2",
                        "section_title": "运营配置",
                        "text": "支持批量配置运费模板。",
                    }
                ],
            },
            ai_analysis={"provider": "DeepSeek", "enabled": False, "risk_table": []},
            token_usage=120,
            cost=0.12,
            duration_ms=800,
        )
        analysis_record = save_analysis_record(
            project_id=project_id,
            code_changes_summary={
                "total_files": 1,
                "total_added": 12,
                "total_removed": 2,
                "files": [
                    {
                        "package": "com.example.order.OrderService",
                        "added": 12,
                        "removed": 2,
                    }
                ],
            },
            test_coverage_result={
                "total_changed_methods": 1,
                "covered": [],
                "uncovered": ["com.example.order.OrderService.createOrder"],
                "coverage_rate": 0.0,
                "details": [
                    {
                        "method": "com.example.order.OrderService.createOrder",
                        "description": "创建订单",
                        "is_covered": False,
                        "matched_tests": [],
                    }
                ],
            },
            test_score=85.0,
            score_snapshot={
                "total_score": 85.0,
                "grade": "B",
                "summary": "妗堜緥鍒嗘瀽瀹屾垚",
                "dimensions": [
                    {
                        "dimension": "姝ラ瀹屾暣鎬?",
                        "score": 80.0,
                        "weight": 0.3,
                        "weighted_score": 24.0,
                        "details": "骞冲潎姝ラ璐ㄩ噺 80.0/100 (5涓敤渚?)",
                    }
                ],
            },
            ai_suggestions={"summary": "案例分析完成"},
            token_usage=180,
            cost=0.18,
            duration_ms=1200,
        )
        return project_id, requirement_record["id"], analysis_record["id"]

    def test_create_list_and_detail_case_quality_record(self, client):
        project_id, requirement_record_id, analysis_record_id = self._prepare_records(client)

        with patch(
            "index.call_deepseek",
            new_callable=AsyncMock,
            return_value={"error": "未配置DEEPSEEK_API_KEY，AI 分析已跳过。", "provider": "DeepSeek"},
        ):
            create_resp = client.post(
                "/api/case-quality/records",
                json={
                    "project_id": project_id,
                    "requirement_analysis_record_id": requirement_record_id,
                    "analysis_record_id": analysis_record_id,
                    "code_changes_file_name": "code-changes.json",
                    "test_cases_file_name": "test-cases.csv",
                },
            )

        assert create_resp.status_code == 200
        create_data = create_resp.json()["data"]
        assert create_data["project_id"] == project_id
        assert create_data["requirement_analysis_record_id"] == requirement_record_id
        assert create_data["analysis_record_id"] == analysis_record_id
        assert create_data["requirement_file_name"] == "requirement.docx"
        assert create_data["total_token_usage"] == 300
        assert create_data["total_duration_ms"] == 2000
        assert create_data["combined_result_snapshot"]["overview"]["project_id"] == project_id

        list_resp = client.get("/api/case-quality/records")
        assert list_resp.status_code == 200
        list_data = list_resp.json()["data"]
        assert len(list_data) == 1
        assert list_data[0]["id"] == create_data["id"]

        detail_resp = client.get(f"/api/case-quality/records/{create_data['id']}")
        assert detail_resp.status_code == 200
        detail_data = detail_resp.json()["data"]
        assert detail_data["requirement_result_snapshot"]["score"]["total_score"] == 72
        assert detail_data["case_result_snapshot"]["score"]["total_score"] == 85.0
        assert detail_data["case_result_snapshot"]["test_case_count"] == 5
        assert detail_data["case_result_snapshot"]["ai_analysis"] is None
        assert detail_data["combined_result_snapshot"]["case_report"]["ai_analysis"] is None

        logs_resp = client.get("/api/audit-logs")
        assert logs_resp.status_code == 200
        logs = logs_resp.json()["data"]
        report_log = next(
            item for item in logs
            if item["action"] == "生成案例质检报告" and item["target_id"] == str(create_data["id"])
        )
        assert report_log["module"] == "功能测试"
        assert report_log["operator_username"] == "admin"
        assert report_log["result"] == "success"

    def test_create_case_quality_record_generates_ai_test_advice(self, client):
        project_id, requirement_record_id, analysis_record_id = self._prepare_records(client)
        mock_ai_response = {
            "provider": "DeepSeek",
            "provider_key": "deepseek",
            "result": {
                "summary": "建议优先围绕需求映射命中点与未覆盖方法补齐核心回归。",
                "overall_assessment": "优先补齐高风险回归",
                "must_test": [
                    {
                        "title": "补测订单提交主链路",
                        "priority": "P0",
                        "reason": "需求点命中核心下单场景，且存在未覆盖变更方法。",
                        "evidence": "RP-1 命中下单流程映射，com.example.order.OrderService.createOrder 当前未覆盖。",
                        "requirement_ids": ["RP-1"],
                        "methods": ["com.example.order.OrderService.createOrder"],
                        "test_focus": "补充库存不足、重复提交和订单回滚断言。",
                        "expected_risk": "下单主链路可能出现库存扣减或幂等校验回归。",
                    }
                ],
                "should_test": [
                    {
                        "title": "补充关联场景回归",
                        "priority": "P1",
                        "reason": "命中需求映射后仍需扩展关联流程验证。",
                        "evidence": "规则建议提示需要补齐库存不足与重复提交验证。",
                        "requirement_ids": ["RP-1"],
                        "methods": [],
                        "test_focus": "覆盖异常分支与边界输入。",
                        "expected_risk": "关联场景遗漏会导致异常流漏测。",
                    }
                ],
                "regression_scope": ["下单流程", "库存校验"],
                "missing_information": ["缺少最近一次线上缺陷数据"],
            },
            "usage": {
                "prompt_tokens": 160,
                "completion_tokens": 96,
                "total_tokens": 256,
            },
        }

        with patch("index.call_deepseek", new_callable=AsyncMock, return_value=mock_ai_response):
            create_resp = client.post(
                "/api/case-quality/records",
                json={
                    "project_id": project_id,
                    "requirement_analysis_record_id": requirement_record_id,
                    "analysis_record_id": analysis_record_id,
                    "code_changes_file_name": "code-changes.json",
                    "test_cases_file_name": "test-cases.csv",
                },
            )

        assert create_resp.status_code == 200
        create_data = create_resp.json()["data"]
        assert create_data["total_token_usage"] == 556
        ai_test_advice = create_data["combined_result_snapshot"]["ai_test_advice"]
        assert ai_test_advice["provider"] == "DeepSeek"
        assert ai_test_advice["enabled"] is True
        assert ai_test_advice["overall_assessment"] == "优先补齐高风险回归"
        assert ai_test_advice["must_test"][0]["priority"] == "P0"
        assert ai_test_advice["must_test"][0]["requirement_ids"] == ["RP-1"]
        assert ai_test_advice["must_test"][0]["methods"] == ["com.example.order.OrderService.createOrder"]
        assert ai_test_advice["regression_scope"] == ["下单流程", "库存校验"]
        assert create_data["case_result_snapshot"]["ai_analysis"] is None
        assert create_data["combined_result_snapshot"]["case_report"]["ai_analysis"] is None

    def test_create_case_quality_record_skips_ai_test_advice_when_use_ai_is_false(self, client):
        project_id, requirement_record_id, analysis_record_id = self._prepare_records(client)

        with patch("index.call_deepseek", new_callable=AsyncMock) as mock_call_deepseek:
            create_resp = client.post(
                "/api/case-quality/records",
                json={
                    "project_id": project_id,
                    "requirement_analysis_record_id": requirement_record_id,
                    "analysis_record_id": analysis_record_id,
                    "code_changes_file_name": "code-changes.json",
                    "test_cases_file_name": "test-cases.csv",
                    "use_ai": False,
                },
            )

        assert create_resp.status_code == 200
        mock_call_deepseek.assert_not_awaited()

        create_data = create_resp.json()["data"]
        assert create_data["total_token_usage"] == 300
        ai_test_advice = create_data["combined_result_snapshot"]["ai_test_advice"]
        assert ai_test_advice["provider"] == "DeepSeek"
        assert ai_test_advice["enabled"] is False
        assert ai_test_advice["must_test"] == []
        assert ai_test_advice["should_test"] == []
        assert ai_test_advice["error"] == "案例质检已关闭 AI，本次不会调用 AI 生成测试建议。"

    def test_create_case_quality_record_rejects_cross_project_records(self, client):
        project_1 = client.post("/api/projects", json={"name": "项目一"}).json()["data"]["id"]
        project_2 = client.post("/api/projects", json={"name": "项目二"}).json()["data"]["id"]

        requirement_record = save_requirement_analysis_record(
            project_id=project_1,
            requirement_file_name="req1.docx",
            section_snapshot={"selected_mode": "preferred_sections", "points": []},
            result_snapshot={"overview": {}, "score": {"total_score": 60}},
            ai_analysis={"risk_table": []},
            token_usage=0,
            cost=0.0,
            duration_ms=0,
        )
        analysis_record = save_analysis_record(
            project_id=project_2,
            code_changes_summary={},
            test_coverage_result={},
            test_score=75.0,
            ai_suggestions=None,
            token_usage=0,
            cost=0.0,
            duration_ms=0,
        )

        resp = client.post(
            "/api/case-quality/records",
            json={
                "project_id": project_1,
                "requirement_analysis_record_id": requirement_record["id"],
                "analysis_record_id": analysis_record["id"],
                "code_changes_file_name": "code-changes.json",
                "test_cases_file_name": "test-cases.csv",
            },
        )

        assert resp.status_code == 400


class TestProjectAnalyze:
    """测试基于项目上下文的分析"""

    def test_analyze_project_not_found(self, client):
        """分析不存在的项目返回404"""
        code_file = FIXTURES_DIR / "sample_code_changes.json"
        test_file = FIXTURES_DIR / "sample_test_cases.csv"
        mapping_file = FIXTURES_DIR / "sample_mapping.csv"

        with open(code_file, "rb") as cf, open(test_file, "rb") as tf, open(mapping_file, "rb") as mf:
            resp = client.post(
                "/api/projects/9999/analyze",
                files={
                    "code_changes": ("code.json", cf, "application/json"),
                    "test_cases_file": ("tests.csv", tf, "text/csv"),
                    "mapping_file": ("mapping.csv", mf, "text/csv"),
                },
                data={"use_ai": "false"},
            )
        assert resp.status_code == 404

    def test_analyze_with_mapping_file(self, client):
        """使用上传的映射文件进行项目分析（不使用AI）"""
        # 创建项目
        create_resp = client.post("/api/projects", json={"name": "分析项目"})
        project_id = create_resp.json()["data"]["id"]

        code_file = FIXTURES_DIR / "sample_code_changes.json"
        test_file = FIXTURES_DIR / "sample_test_cases.csv"
        mapping_file = FIXTURES_DIR / "sample_mapping.csv"

        with open(code_file, "rb") as cf, open(test_file, "rb") as tf, open(mapping_file, "rb") as mf:
            resp = client.post(
                f"/api/projects/{project_id}/analyze",
                files={
                    "code_changes": ("code.json", cf, "application/json"),
                    "test_cases_file": ("tests.csv", tf, "text/csv"),
                    "mapping_file": ("mapping.csv", mf, "text/csv"),
                },
                data={"use_ai": "false"},
            )
        assert resp.status_code == 200
        data = resp.json()
        assert data["success"] is True
        assert "record_id" in data["data"]
        assert "diff_analysis" in data["data"]
        assert "coverage" in data["data"]
        assert "score" in data["data"]
        assert data["data"]["test_case_count"] == 4

        # 验证记录已保存
        record_id = data["data"]["record_id"]
        record_resp = client.get(f"/api/records/{record_id}")
        assert record_resp.status_code == 200
        assert record_resp.json()["data"]["test_case_count"] == 4

        logs_resp = client.get("/api/audit-logs")
        assert logs_resp.status_code == 200
        logs = logs_resp.json()["data"]
        analyze_log = next(
            item for item in logs
            if item["action"] == "案例分析" and item["target_id"] == str(record_id)
        )
        assert analyze_log["module"] == "功能测试"
        assert analyze_log["operator_username"] == "admin"
        assert analyze_log["result"] == "success"

    def test_analyze_with_stored_mapping(self, client):
        """使用项目存储的映射数据进行分析"""
        # 创建项目并上传映射
        create_resp = client.post("/api/projects", json={"name": "有映射的项目"})
        project_id = create_resp.json()["data"]["id"]

        mapping_file = FIXTURES_DIR / "sample_mapping.csv"
        with open(mapping_file, "rb") as mf:
            client.post(
                f"/api/projects/{project_id}/mapping",
                files={"mapping_file": ("mapping.csv", mf, "text/csv")},
            )

        # 不提供映射文件进行分析
        code_file = FIXTURES_DIR / "sample_code_changes.json"
        test_file = FIXTURES_DIR / "sample_test_cases.csv"

        with open(code_file, "rb") as cf, open(test_file, "rb") as tf:
            resp = client.post(
                f"/api/projects/{project_id}/analyze",
                files={
                    "code_changes": ("code.json", cf, "application/json"),
                    "test_cases_file": ("tests.csv", tf, "text/csv"),
                },
                data={"use_ai": "false"},
            )
        assert resp.status_code == 200
        assert resp.json()["success"] is True

    def test_analyze_with_real_template_excel(self, client):
        """支持带说明首行和第二行表头的真实 Excel 测试用例模板"""
        create_resp = client.post("/api/projects", json={"name": "真实用例模板项目"})
        project_id = create_resp.json()["data"]["id"]

        code_file = FIXTURES_DIR / "sample_code_changes.json"
        mapping_file = FIXTURES_DIR / "sample_mapping.csv"
        test_file_bytes = build_real_test_cases_xlsx_bytes()

        with open(code_file, "rb") as cf, open(mapping_file, "rb") as mf:
            resp = client.post(
                f"/api/projects/{project_id}/analyze",
                files={
                    "code_changes": ("code.json", cf, "application/json"),
                    "test_cases_file": (
                        "real-test-cases.xlsx",
                        test_file_bytes,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    ),
                    "mapping_file": ("mapping.csv", mf, "text/csv"),
                },
                data={"use_ai": "false"},
            )

        assert resp.status_code == 200
        data = resp.json()
        assert data["success"] is True
        assert data["data"]["coverage"]["total_changed_methods"] > 0

        details = data["data"]["coverage"]["details"]
        delete_user_detail = next(
            detail for detail in details
            if detail["method"] == "com.example.user.UserService.deleteUser"
        )
        calculate_total_detail = next(
            detail for detail in details
            if detail["method"] == "com.example.order.OrderService.calculateTotal"
        )

        assert delete_user_detail["matched_tests"] == ["case001"]
        assert calculate_total_detail["matched_tests"] == ["case002"]

    def test_analyze_with_line_array_code_json(self, client, sample_code_changes_dict):
        """代码改动 JSON 支持按行数组格式上传"""
        create_resp = client.post("/api/projects", json={"name": "逐行代码项目"})
        project_id = create_resp.json()["data"]["id"]

        mapping_file = FIXTURES_DIR / "sample_mapping.csv"
        test_file = FIXTURES_DIR / "sample_test_cases.csv"

        code_payload = json.loads(json.dumps(sample_code_changes_dict))
        for field in ("current", "history"):
            code_payload["data"][field] = [item.split("\n") for item in code_payload["data"][field]]
        code_bytes = json.dumps(code_payload, ensure_ascii=False).encode("utf-8")

        with open(mapping_file, "rb") as mf, open(test_file, "rb") as tf:
            resp = client.post(
                f"/api/projects/{project_id}/analyze",
                files={
                    "code_changes": ("code.json", code_bytes, "application/json"),
                    "test_cases_file": ("tests.csv", tf, "text/csv"),
                    "mapping_file": ("mapping.csv", mf, "text/csv"),
                },
                data={"use_ai": "false"},
            )

        assert resp.status_code == 200
        assert resp.json()["success"] is True

    def test_analyze_with_flat_line_sequence_code_json(self, client, sample_code_changes_dict):
        """兼容误传的顶层逐行数组格式"""
        create_resp = client.post("/api/projects", json={"name": "误传逐行代码项目"})
        project_id = create_resp.json()["data"]["id"]

        mapping_file = FIXTURES_DIR / "sample_mapping.csv"
        test_file = FIXTURES_DIR / "sample_test_cases.csv"

        code_payload = json.loads(json.dumps(sample_code_changes_dict))
        for field in ("current", "history"):
            code_payload["data"][field] = code_payload["data"][field][0].split("\n")
        code_bytes = json.dumps(code_payload, ensure_ascii=False).encode("utf-8")

        with open(mapping_file, "rb") as mf, open(test_file, "rb") as tf:
            resp = client.post(
                f"/api/projects/{project_id}/analyze",
                files={
                    "code_changes": ("code.json", code_bytes, "application/json"),
                    "test_cases_file": ("tests.csv", tf, "text/csv"),
                    "mapping_file": ("mapping.csv", mf, "text/csv"),
                },
                data={"use_ai": "false"},
            )

        assert resp.status_code == 200
        assert resp.json()["success"] is True

    def test_analyze_no_mapping_available(self, client):
        """项目无映射且未上传映射文件应返回400"""
        create_resp = client.post("/api/projects", json={"name": "无映射项目"})
        project_id = create_resp.json()["data"]["id"]

        code_file = FIXTURES_DIR / "sample_code_changes.json"
        test_file = FIXTURES_DIR / "sample_test_cases.csv"

        with open(code_file, "rb") as cf, open(test_file, "rb") as tf:
            resp = client.post(
                f"/api/projects/{project_id}/analyze",
                files={
                    "code_changes": ("code.json", cf, "application/json"),
                    "test_cases_file": ("tests.csv", tf, "text/csv"),
                },
                data={"use_ai": "false"},
            )
        assert resp.status_code == 400

    def test_analyze_with_ai_mock(self, client):
        """使用mock的AI进行项目分析"""
        # 创建项目
        create_resp = client.post("/api/projects", json={"name": "AI分析项目"})
        project_id = create_resp.json()["data"]["id"]

        # Mock DeepSeek
        mock_ai_response = {
            "result": {
                "uncovered_methods": [],
                "coverage_gaps": "无明显缺口",
                "suggested_test_cases": [],
                "risk_assessment": "low",
                "improvement_suggestions": ["增加边界测试"],
            },
            "usage": {
                "prompt_tokens": 500,
                "completion_tokens": 200,
                "total_tokens": 700,
                "prompt_cache_hit_tokens": 100,
                "prompt_cache_miss_tokens": 400,
            },
        }

        code_file = FIXTURES_DIR / "sample_code_changes.json"
        test_file = FIXTURES_DIR / "sample_test_cases.csv"
        mapping_file = FIXTURES_DIR / "sample_mapping.csv"

        with patch("index.call_deepseek", new_callable=AsyncMock, return_value=mock_ai_response):
            with open(code_file, "rb") as cf, open(test_file, "rb") as tf, open(mapping_file, "rb") as mf:
                resp = client.post(
                    f"/api/projects/{project_id}/analyze",
                    files={
                        "code_changes": ("code.json", cf, "application/json"),
                        "test_cases_file": ("tests.csv", tf, "text/csv"),
                        "mapping_file": ("mapping.csv", mf, "text/csv"),
                    },
                    data={"use_ai": "true"},
                )

        assert resp.status_code == 200
        data = resp.json()
        assert data["success"] is True
        assert data["data"]["ai_analysis"] is not None
        assert data["data"]["ai_cost"] is not None
        assert "record_id" in data["data"]


class TestIssueAnalysis:
    """测试问题归纳分析上传接口"""

    @staticmethod
    def _build_issue_excel_bytes() -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(
            [
                "出现该问题的原因",
                "改善举措",
                "发生阶段",
                "是否人为原因",
                "发生原因总结",
                "标签",
            ]
        )
        sheet.append(
            [
                "需求评审不足，边界条件遗漏",
                "补充评审清单；增加边界场景检查",
                "需求阶段",
                "是",
                "需求澄清不足",
                "需求,边界场景",
            ]
        )
        sheet.append(
            [
                "联调环境不稳定",
                "稳定测试环境",
                "联调阶段",
                "否",
                "环境问题",
                "环境",
            ]
        )

        content = io.BytesIO()
        workbook.save(content)
        workbook.close()
        return content.getvalue()

    def test_import_issue_analysis_excel(self, client):
        """上传 Excel 并返回归纳图表数据"""
        resp = client.post(
            "/api/issue-analysis/import",
            files={
                "file": (
                    "issue-analysis.xlsx",
                    TestIssueAnalysis._build_issue_excel_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert resp.status_code == 200
        data = resp.json()
        assert data["success"] is True
        assert data["data"]["overview"]["total_records"] == 2
        assert data["data"]["charts"]["stage_distribution"]
        assert data["data"]["charts"]["human_factor_distribution"]
        assert data["data"]["summary"]["key_findings"]

    def test_import_issue_analysis_rejects_missing_fields(self, client):
        """缺少必要字段时返回 400"""
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["问题原因", "阶段"])
        sheet.append(["测试遗漏", "测试阶段"])

        content = io.BytesIO()
        workbook.save(content)
        workbook.close()

        resp = client.post(
            "/api/issue-analysis/import",
            files={
                "file": (
                    "invalid-issue-analysis.xlsx",
                    content.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert resp.status_code == 400
        assert "缺少必要字段" in resp.json()["detail"]


class TestTestIssueFiles:
    """测试问题文件上传接口"""

    @staticmethod
    def _build_test_issue_excel_bytes() -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(DEFECT_FIELDS)
        sheet.append(
            [
                "BUG-001",
                "登录接口返回空指针",
                "TASK-001",
                "智测平台",
                "ZCPT",
                "REQ-001",
                "2026-03-01",
                "已关闭",
                "张三",
                "zhangsan",
                "严重",
                "必现",
                "影响核心交易",
                "系统测试",
                "接口校验缺失",
                "边界值遗漏",
                "请求参数为空时接口抛出异常",
                "补充空值判断",
                "系统测试",
                "李四",
                "lisi",
                "8",
                "1",
                "账户中心",
                "2026-03-03 12:00:00",
                "开发一组",
                "测试一组",
                "核心交易回归",
                "登录模块",
                "登录接口",
                "王五",
                "wangwu",
                "2026-03-02 09:00:00",
                "否",
                "",
            ]
        )

        content = io.BytesIO()
        workbook.save(content)
        workbook.close()
        return content.getvalue()

    def test_upload_test_issue_file_and_list(self, client):
        """上传测试问题文件并绑定项目"""
        create_resp = client.post("/api/projects", json={"name": "核心项目", "description": "项目描述"})
        project_id = create_resp.json()["data"]["id"]

        upload_resp = client.post(
            "/api/test-issue-files",
            data={"project_id": str(project_id)},
            files={
                "file": (
                    "test-issues.xlsx",
                    self._build_test_issue_excel_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert upload_resp.status_code == 200
        upload_data = upload_resp.json()
        assert upload_data["success"] is True
        assert upload_data["data"]["project_id"] == project_id
        assert upload_data["data"]["project_name"] == "核心项目"
        assert upload_data["data"]["row_count"] == 1

        list_resp = client.get("/api/test-issue-files", params={"project_id": project_id})
        assert list_resp.status_code == 200
        list_data = list_resp.json()
        assert list_data["success"] is True
        assert len(list_data["data"]) == 1
        assert list_data["data"][0]["project_name"] == "核心项目"

        analysis_resp = client.get(f"/api/test-issue-files/{upload_data['data']['id']}/analysis")
        assert analysis_resp.status_code == 200
        analysis_data = analysis_resp.json()
        assert analysis_data["success"] is True
        assert analysis_data["data"]["overview"]["total_records"] == 1

    def test_upload_test_issue_file_rejects_unknown_project(self, client):
        """绑定不存在项目时返回 404"""
        resp = client.post(
            "/api/test-issue-files",
            data={"project_id": "9999"},
            files={
                "file": (
                    "test-issues.xlsx",
                    self._build_test_issue_excel_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert resp.status_code == 404
        assert "项目不存在" in resp.json()["detail"]

    def test_upload_production_issue_file_and_list(self, client):
        """上传生产问题文件并在列表中可见"""
        upload_resp = client.post(
            "/api/production-issue-files",
            files={
                "file": (
                    "issue-analysis.xlsx",
                    TestIssueAnalysis._build_issue_excel_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert upload_resp.status_code == 200
        upload_data = upload_resp.json()
        assert upload_data["success"] is True
        assert upload_data["data"]["file_name"] == "issue-analysis.xlsx"
        assert upload_data["data"]["file_type"] == "excel"
        assert upload_data["data"]["row_count"] == 2

        list_resp = client.get("/api/production-issue-files")
        assert list_resp.status_code == 200
        list_data = list_resp.json()
        assert list_data["success"] is True
        assert len(list_data["data"]) == 1
        assert list_data["data"][0]["id"] == upload_data["data"]["id"]

        analysis_resp = client.get(f"/api/production-issue-files/{upload_data['data']['id']}/analysis")
        assert analysis_resp.status_code == 200
        analysis_data = analysis_resp.json()
        assert analysis_data["success"] is True
        assert analysis_data["data"]["overview"]["total_records"] == 2

    def test_upload_production_issue_file_rejects_missing_fields(self, client):
        """上传文件缺少必要字段时返回 400"""
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["问题原因", "阶段"])
        sheet.append(["测试遗漏", "测试阶段"])

        content = io.BytesIO()
        workbook.save(content)
        workbook.close()

        resp = client.post(
            "/api/production-issue-files",
            files={
                "file": (
                    "invalid-issue-analysis.xlsx",
                    content.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert resp.status_code == 400
        assert "缺少必要字段" in resp.json()["detail"]


class TestRequirementMappings:
    @staticmethod
    def _build_requirement_mapping_xlsx_bytes() -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet.append(["标签", "需求关键字", "关联场景"])
        sheet.merge_cells("A2:A3")
        sheet.merge_cells("B2:B3")
        sheet["A2"] = "流程变更"
        sheet["B2"] = "抄录"
        sheet["C2"] = "一键抄录"
        sheet["C3"] = "逐字抄录"

        buffer = io.BytesIO()
        workbook.save(buffer)
        workbook.close()
        return buffer.getvalue()

    @staticmethod
    def _build_requirement_mapping_xls_bytes() -> bytes:
        import xlwt

        workbook = xlwt.Workbook()
        sheet = workbook.add_sheet("Sheet1")
        sheet.write(0, 0, "标签")
        sheet.write(0, 1, "需求关键字")
        sheet.write(0, 2, "关联场景")
        sheet.write_merge(1, 2, 0, 0, "流程变更")
        sheet.write_merge(1, 2, 1, 1, "抄录")
        sheet.write(1, 2, "一键抄录")
        sheet.write(2, 2, "逐字抄录")

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def test_upload_and_get_requirement_mapping_xlsx(self, client):
        create_resp = client.post("/api/projects", json={"name": "需求映射项目"})
        project_id = create_resp.json()["data"]["id"]

        upload_resp = client.post(
            f"/api/projects/{project_id}/requirement-mapping",
            files={
                "file": (
                    "requirement-mapping.xlsx",
                    self._build_requirement_mapping_xlsx_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert upload_resp.status_code == 200
        upload_data = upload_resp.json()["data"]
        assert upload_data["source_type"] == "upload"
        assert upload_data["group_count"] == 1
        assert upload_data["row_count"] == 2
        assert upload_data["last_file_type"] == "xlsx"
        assert upload_data["rows"][0]["tag_row_span"] == 2

        detail_resp = client.get(f"/api/projects/{project_id}/requirement-mapping")
        assert detail_resp.status_code == 200
        detail_data = detail_resp.json()["data"]
        assert detail_data["groups"][0]["related_scenarios"] == ["一键抄录", "逐字抄录"]

    def test_upload_requirement_mapping_xls(self, client):
        create_resp = client.post("/api/projects", json={"name": "xls项目"})
        project_id = create_resp.json()["data"]["id"]

        resp = client.post(
            f"/api/projects/{project_id}/requirement-mapping",
            files={
                "file": (
                    "requirement-mapping.xls",
                    self._build_requirement_mapping_xls_bytes(),
                    "application/vnd.ms-excel",
                )
            },
        )

        assert resp.status_code == 200
        assert resp.json()["data"]["last_file_type"] == "xls"

    def test_manual_put_sets_manual_or_mixed_source(self, client):
        create_resp = client.post("/api/projects", json={"name": "手工维护项目"})
        project_id = create_resp.json()["data"]["id"]

        manual_resp = client.put(
            f"/api/projects/{project_id}/requirement-mapping",
            json={
                "groups": [
                    {
                        "id": "manual-1",
                        "tag": "页面新增",
                        "requirement_keyword": "新增页面",
                        "related_scenarios": ["兼容性测试", "跳转链路"],
                    }
                ]
            },
        )

        assert manual_resp.status_code == 200
        assert manual_resp.json()["data"]["source_type"] == "manual"

        upload_resp = client.post(
            f"/api/projects/{project_id}/requirement-mapping",
            files={
                "file": (
                    "requirement-mapping.xlsx",
                    self._build_requirement_mapping_xlsx_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert upload_resp.status_code == 200
        assert upload_resp.json()["data"]["source_type"] == "upload"

        mixed_resp = client.put(
            f"/api/projects/{project_id}/requirement-mapping",
            json={
                "groups": [
                    {
                        "id": "group-1",
                        "tag": "流程变更",
                        "requirement_keyword": "抄录",
                        "related_scenarios": ["一键抄录", "逐字抄录", "逐字点击"],
                    }
                ]
            },
        )

        assert mixed_resp.status_code == 200
        assert mixed_resp.json()["data"]["source_type"] == "mixed"
        assert mixed_resp.json()["data"]["last_file_name"] == "requirement-mapping.xlsx"

    def test_put_empty_groups_deletes_current_mapping(self, client):
        create_resp = client.post("/api/projects", json={"name": "删除项目"})
        project_id = create_resp.json()["data"]["id"]
        client.put(
            f"/api/projects/{project_id}/requirement-mapping",
            json={
                "groups": [
                    {
                        "id": "manual-1",
                        "tag": "页面新增",
                        "requirement_keyword": "新增页面",
                        "related_scenarios": ["兼容性测试"],
                    }
                ]
            },
        )

        delete_resp = client.put(
            f"/api/projects/{project_id}/requirement-mapping",
            json={"groups": []},
        )

        assert delete_resp.status_code == 200
        assert delete_resp.json()["data"] is None

        detail_resp = client.get(f"/api/projects/{project_id}/requirement-mapping")
        assert detail_resp.status_code == 404

    def test_download_requirement_mapping_template(self, client):
        resp = client.get("/api/requirement-mapping-template")

        assert resp.status_code == 200
        assert resp.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "attachment; filename=\"requirement-mapping-template.xlsx\"" in resp.headers["content-disposition"]

    def test_reupload_replaces_manual_data(self, client):
        create_resp = client.post("/api/projects", json={"name": "覆盖项目"})
        project_id = create_resp.json()["data"]["id"]

        client.put(
            f"/api/projects/{project_id}/requirement-mapping",
            json={
                "groups": [
                    {
                        "id": "manual-1",
                        "tag": "弹窗",
                        "requirement_keyword": "新增弹窗",
                        "related_scenarios": ["弹窗核对"],
                    }
                ]
            },
        )

        upload_resp = client.post(
            f"/api/projects/{project_id}/requirement-mapping",
            files={
                "file": (
                    "requirement-mapping.xlsx",
                    self._build_requirement_mapping_xlsx_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert upload_resp.status_code == 200
        data = upload_resp.json()["data"]
        assert data["source_type"] == "upload"
        assert data["groups"][0]["tag"] == "流程变更"
        assert data["groups"][0]["related_scenarios"] == ["一键抄录", "逐字抄录"]
