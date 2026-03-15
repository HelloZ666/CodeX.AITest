"""
test_file_upload.py - 文件上传和解析测试
"""

import io
import json
import re
import zipfile

import pytest
from openpyxl import Workbook

from services.file_parser import (
    detect_file_type,
    parse_csv,
    parse_excel,
    parse_json,
    validate_file,
)


def build_excel_bytes(sheet_rows: list[tuple[str, list[list[object]]]]) -> bytes:
    workbook = Workbook()
    default_sheet = workbook.active

    for index, (title, rows) in enumerate(sheet_rows):
        worksheet = default_sheet if index == 0 else workbook.create_sheet(title=title)
        worksheet.title = title
        for row in rows:
            worksheet.append(row)

    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def override_sheet_dimension(content: bytes, dimension_ref: str) -> bytes:
    source = io.BytesIO(content)
    target = io.BytesIO()

    with zipfile.ZipFile(source, "r") as source_zip, zipfile.ZipFile(target, "w", zipfile.ZIP_DEFLATED) as target_zip:
        for item in source_zip.infolist():
            data = source_zip.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                xml = data.decode("utf-8")
                xml = re.sub(r'<dimension ref="[^"]*"\s*/>', f'<dimension ref="{dimension_ref}"/>', xml, count=1)
                data = xml.encode("utf-8")
            target_zip.writestr(item, data)

    return target.getvalue()


class TestParseCSV:
    """测试 CSV 解析"""

    def test_parse_normal_csv(self, sample_mapping_csv):
        rows = parse_csv(sample_mapping_csv)
        assert len(rows) > 0
        assert "包名" in rows[0]

    def test_parse_bytes_utf8(self, sample_mapping_csv):
        content = sample_mapping_csv.encode("utf-8")
        rows = parse_csv(content)
        assert len(rows) > 0

    def test_parse_bytes_utf8_bom(self, sample_mapping_csv):
        content = b"\xef\xbb\xbf" + sample_mapping_csv.encode("utf-8")
        rows = parse_csv(content)
        assert len(rows) > 0

    def test_parse_empty_csv(self):
        with pytest.raises(ValueError, match="内容为空"):
            parse_csv("")

    def test_parse_header_only(self):
        with pytest.raises(ValueError, match="没有数据行"):
            parse_csv("col1,col2,col3\n")


class TestParseExcel:
    """测试 Excel 解析"""

    def test_parse_xls_content(self):
        import xlwt

        workbook = xlwt.Workbook()
        sheet = workbook.add_sheet("Sheet1")
        sheet.write(0, 0, "标签")
        sheet.write(0, 1, "需求关键字")
        sheet.write(0, 2, "关联场景")
        sheet.write(1, 0, "流程变更")
        sheet.write(1, 1, "抄录")
        sheet.write(1, 2, "一键抄录")

        buffer = io.BytesIO()
        workbook.save(buffer)

        rows = parse_excel(buffer.getvalue())

        assert rows == [{"标签": "流程变更", "需求关键字": "抄录", "关联场景": "一键抄录"}]

    def test_parse_excel_skips_leading_empty_rows(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "问题归纳"
        worksheet["A3"] = "原因"
        worksheet["B3"] = "措施"
        worksheet["A4"] = "配置错误"
        worksheet["B4"] = "补充校验"

        buffer = io.BytesIO()
        workbook.save(buffer)
        workbook.close()

        rows = parse_excel(buffer.getvalue())

        assert rows == [{"原因": "配置错误", "措施": "补充校验"}]

    def test_parse_excel_uses_next_sheet_when_active_sheet_is_empty(self):
        content = build_excel_bytes(
            [
                ("空白sheet", []),
                ("缺陷总结", [["缺陷原因", "缺陷来源"], ["空指针", "系统测试"]]),
            ]
        )

        rows = parse_excel(content)

        assert rows == [{"缺陷原因": "空指针", "缺陷来源": "系统测试"}]

    def test_parse_excel_header_only(self):
        content = build_excel_bytes([("Sheet1", [["col1", "col2"]])])

        with pytest.raises(ValueError, match="没有数据行"):
            parse_excel(content)

    def test_parse_excel_ignores_invalid_sheet_dimension_metadata(self):
        content = build_excel_bytes(
            [("缺陷清单", [["缺陷ID", "缺陷摘要"], ["B2026001", "登录页报错"], ["B2026002", "首页按钮缺失"]])]
        )
        content = override_sheet_dimension(content, "A1:A1")

        rows = parse_excel(content)

        assert len(rows) == 2
        assert rows[0]["缺陷ID"] == "B2026001"
        assert rows[1]["缺陷摘要"] == "首页按钮缺失"


class TestParseJSON:
    """测试 JSON 解析"""

    def test_parse_normal_json(self, sample_code_changes_json):
        result = parse_json(sample_code_changes_json)
        assert isinstance(result, dict)
        assert "success" in result

    def test_parse_bytes(self, sample_code_changes_json):
        content = sample_code_changes_json.encode("utf-8")
        result = parse_json(content)
        assert isinstance(result, dict)

    def test_parse_empty(self):
        with pytest.raises(ValueError, match="内容为空"):
            parse_json("")

    def test_parse_invalid(self):
        with pytest.raises(ValueError, match="JSON格式无效"):
            parse_json("{invalid json")


class TestDetectFileType:
    """测试文件类型检测"""

    def test_csv(self):
        assert detect_file_type("test.csv") == "csv"

    def test_excel_xlsx(self):
        assert detect_file_type("test.xlsx") == "excel"

    def test_excel_xls(self):
        assert detect_file_type("test.xls") == "excel"

    def test_json(self):
        assert detect_file_type("data.json") == "json"

    def test_unknown(self):
        assert detect_file_type("file.txt") == "unknown"

    def test_case_insensitive(self):
        assert detect_file_type("TEST.CSV") == "csv"
        assert detect_file_type("DATA.JSON") == "json"


class TestValidateFile:
    """测试文件校验"""

    def test_valid_csv(self):
        err = validate_file("test.csv", b"content", ["csv"])
        assert err == ""

    def test_valid_json(self):
        err = validate_file("data.json", b"content", ["json"])
        assert err == ""

    def test_unsupported_type(self):
        err = validate_file("file.txt", b"content", ["csv", "json"])
        assert "不支持" in err

    def test_wrong_type_for_endpoint(self):
        err = validate_file("test.csv", b"content", ["json"])
        assert "不支持" in err

    def test_file_too_large(self):
        large_content = b"x" * (11 * 1024 * 1024)
        err = validate_file("test.csv", large_content, ["csv"], max_size_mb=10.0)
        assert "过大" in err

    def test_file_within_size(self):
        content = b"x" * 1024
        err = validate_file("test.csv", content, ["csv"], max_size_mb=10.0)
        assert err == ""
