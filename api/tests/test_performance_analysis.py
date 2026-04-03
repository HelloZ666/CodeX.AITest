import io
from datetime import date

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

from services.database import init_db
from services.performance_analysis import analyze_performance_workbook


@pytest.fixture(autouse=True)
def temp_db(tmp_path, monkeypatch):
    db_path = str(tmp_path / "test_performance_analysis.db")
    monkeypatch.setattr("services.database.get_db_path", lambda: db_path)
    monkeypatch.setattr("services.performance_analysis_file_store.get_db_path", lambda: db_path)
    init_db()
    return db_path


@pytest.fixture
def client():
    from index import app

    with TestClient(app) as test_client:
        login_resp = test_client.post(
            "/api/auth/login",
            json={"username": "admin", "password": "Admin123!"},
        )
        assert login_resp.status_code == 200
        yield test_client


def _append_history_sheet(workbook: Workbook, name: str, summary_rows: list[list[object]], per_capita_rows: list[list[object]]):
    sheet = workbook.create_sheet(name)
    sheet.append(["一、历年汇总数据"])
    sheet.append([
        "月份",
        "同步任务数",
        "同步+回归",
        "发布总次数",
        "需求数(同步+需求号去重)",
        "缺陷数 除性能、代码扫描外所有任务类型（SIT+FT）",
        None,
        "测试任务平均时效",
        "设计用例数",
        "执行案例数",
        "功能人月投入",
        "性能人月投入",
        "QA人月投入",
    ])
    sheet.merge_cells(start_row=2, start_column=6, end_row=2, end_column=7)
    for row in summary_rows:
        sheet.append([
            row[0],
            row[1],
            row[2],
            row[3],
            row[4],
            row[5],
            None,
            row[6],
            row[7],
            row[8],
            row[9],
            row[10],
            row[11],
        ])

    sheet.append(["2.月度人均任务数据"])
    sheet.append([
        "月份",
        "人均同步任务",
        "人均（同步+回归）",
        "人均需求数",
        "人均总缺陷数",
        "缺陷率",
        None,
        "平均设计案例数",
        "执行案例数（同步+回归+安全）",
    ])
    sheet.merge_cells(start_row=7, start_column=6, end_row=7, end_column=7)
    for row in per_capita_rows:
        sheet.append([
            row[0],
            row[1],
            row[2],
            row[3],
            row[4],
            row[5],
            None,
            row[6],
            row[7],
        ])


def _append_current_summary_sheet(workbook: Workbook, name: str, rows: list[list[object]]):
    sheet = workbook.create_sheet(name)
    sheet.append(["一、当年月度汇总数据"])
    sheet.append([
        "月份",
        "同步任务数",
        "同步+回归",
        "发布总次数",
        "需求数(同步+需求号去重)",
        "缺陷数",
        "测试任务平均时效",
        "设计用例数",
        "执行案例数",
        "功能人月投入",
        "性能人月投入",
        "QA人月投入",
    ])
    for row in rows:
        sheet.append(row)
    sheet.append(["2.月度人均任务数据"])


def _append_current_external_sheet(workbook: Workbook, name: str, business: str, rows: list[list[object]]):
    sheet = workbook.create_sheet(name)
    sheet.append(["月份&人力", None, None, "测试任务&测试案例", None, None, None, None, None, None, "缺陷信息"])
    sheet.append([
        "月份",
        None,
        "人月投入",
        "发布总次数",
        "同步任务数",
        "需求数",
        "设计案例数",
        "执行案例数",
        "缺陷率",
        "测试任务平均时效",
        "测试缺陷数",
        "生产缺陷数",
        "生产缺陷检出率",
        "自动化覆盖率",
        "自动化执行通过率",
        "计划应用数",
        "已接入应用数",
        "精准接入率",
    ])
    for row in rows:
        sheet.append([row[0], business, *row[1:]])


def _append_current_team_sheet(workbook: Workbook, name: str, month_rows: dict[int, list[list[object]]]):
    sheet = workbook.create_sheet(name)
    for month, rows in month_rows.items():
        sheet.append([None] * 21)
        sheet.append([f"2026年{month}月份"] + [None] * 20)
        sheet.append([
            "团队",
            "系统个数",
            "同步任务数",
            "同步任务数+\n回归任务数",
            "总需求数",
            "总BUG数\n（全任务）",
            "设计案例数\n（同步任务）",
            "执行案例数\n（同步+回归(手工&自动化)+安全）",
            "人月数",
            "人均任务",
            None,
            "人均需求数",
            None,
            "人均缺陷数",
            None,
            "缺陷率",
            None,
            "平均设计案例数",
            None,
            "平均执行案例数",
            None,
        ])
        sheet.append([None, None, None, None, None, None, None, None, None, "数量", "排名", "数量", "排名", "数量", "排名", "数量", "排名", "数量", "排名", "数量", "排名"])
        for row in rows:
            sheet.append(row)


def _append_distractor_sheets(workbook: Workbook):
    sheet = workbook.create_sheet("寿险对外数据（2025）")
    sheet.append(["月份&人力"])
    sheet.append(["月份", None, "人月投入", "发布总次数", "同步任务数", "需求数", "设计案例数", "执行案例数", "缺陷率", "测试任务平均时效", "测试缺陷数"])
    sheet.append(["1月", "寿险", 999, 999, 9999, 9999, 99999, 99999, 0.9, 99, 999])

    workbook.create_sheet("指标口径定义").append(["本表不参与计算"])
    workbook.create_sheet("各团队数据（寿险&健康险）-2025").append(["混合表不参与计算"])


def _build_workbook_bytes() -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)

    _append_history_sheet(
        workbook,
        "寿险汇总数据-历年数据",
        [
            ["2023年平均数据", 866.08, 1124.5, 258.42, 692.33, 1380.17, 4.75, 807913.5, 1343149.83, 144.4, None, 4],
            ["2024年平均数据", 942.5, 1262.5, 315.92, 731.92, 2432.17, 4.41, 1752366.75, 2590632.17, 114.59, 7.67, 3.67],
            ["2025年平均数据", 1136.42, 1443.67, 307.58, 878.5, 3041, 4.79, 2454535.25, 3444162.58, 116.58, 8.08, 4],
        ],
        [
            ["23年人均任务数", 6, 7.79, 4.79, 9.56, 0.00171, 5594.97, 9301.59],
            ["24年人均任务数", 8.22, 11.02, 6.39, 21.22, 0.00139, 15292.49, 22607.84],
            ["25年人均任务数", 9.75, 12.38, 7.54, 26.09, 0.00124, 21054.51, 29543.34],
        ],
    )
    _append_history_sheet(
        workbook,
        "健康险汇总数据-历年数据",
        [
            ["2024年平均数据", 108.5, 157.67, 49.17, 95.92, 37.67, 3.98, 37882.25, 68710, 9.59, None, None],
            ["2025年平均数据", 135.75, 193.42, 57.5, 117.75, 103.42, 4.56, 71890.75, 106039.25, 11.45, None, None],
        ],
        [
            ["24年人均任务数", 11.31, 16.44, 10.0, 3.93, 0.00099, 3950.18, 7164.75],
            ["25年人均任务数", 11.86, 16.89, 10.28, 9.03, 0.00144, 6278.67, 9261.07],
        ],
    )

    _append_current_summary_sheet(
        workbook,
        "寿险汇总数据-含外协（2026）",
        [
            ["2026年1月数据", 1061, 1153, 305, 836, 1995, 4.34, 2116264, 3181423, 121.6, 1, 1],
            ["2026年2月数据", 480, 606, 126, 387, 986, 5.2, 861716, 1266448, 118.5, 1, 1],
        ],
    )
    _append_current_summary_sheet(
        workbook,
        "健康险汇总数据-含外协（2026）",
        [
            [1, 131, 158, 60, 114, 163, 4.53, 77258, 120106, 13.75, None, None],
            [2, 61, 83, 22, 45, 44, 3.97, 36416, 53472, 13.75, None, None],
        ],
    )

    _append_current_external_sheet(
        workbook,
        "寿险对外数据（2026）",
        "寿险",
        [
            ["1月", 121.6, 305, 1061, 836, 2116264, 3181423, 0.000942699, 4.34, 1995, 1, 0.9, 0.7, 0.95, 10, 8, 0.8],
            ["2月", 118.5, 126, 480, 387, 861716, 1266448, 0.001144228, 5.2, 986, 0, 1, 0.72, 0.96, 10, 9, 0.9],
        ],
    )
    _append_current_external_sheet(
        workbook,
        "健康险对外数据（2026）",
        "健康险",
        [
            ["1月", 13.75, 60, 131, 114, 77258, 120106, 0.002109814, 4.53, 163, 0, 1, None, None, 4, 3, 0.75],
            ["2月", 13.75, 22, 61, 45, 36416, 53472, 0.00120826, 3.97, 44, 0, 1, None, None, 4, 4, 1],
        ],
    )

    _append_current_team_sheet(
        workbook,
        "各团队数据（寿险）-2026",
        {
            1: [["A团队", 12, 60, 90, 45, 20, 6000, 9800, 6, 15, 1, 7.5, 1, 3.33, 2, 0.0033, 2, 1000, 2, 1633.33, 2]],
            2: [["A团队", 12, 70, 100, 50, 22, 6800, 11000, 6.5, 15.38, 1, 7.69, 1, 3.38, 2, 0.0032, 1, 1046.15, 2, 1692.31, 2]],
        },
    )
    _append_current_team_sheet(
        workbook,
        "各团队数据（健康险）-2026",
        {
            1: [["H团队", 6, 30, 45, 24, 8, 2500, 4200, 3.2, 14.06, 1, 7.5, 1, 2.5, 1, 0.0032, 1, 781.25, 1, 1312.5, 1]],
            2: [["H团队", 6, 18, 25, 14, 6, 1800, 3000, 2.8, 8.93, 2, 5, 2, 3.57, 2, 0.0056, 2, 642.86, 2, 1071.43, 2]],
        },
    )

    _append_distractor_sheets(workbook)

    content = io.BytesIO()
    workbook.save(content)
    workbook.close()
    return content.getvalue()


def _build_split_current_summary_workbook_bytes() -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)

    _append_current_summary_sheet(
        workbook,
        "寿险汇总数据-含外协（2026）",
        [
            [1, 1061, 1153, 305, 836, 1995, 4.34, 2116264, 3181423, 121.6, 1, 1],
        ],
    )
    _append_current_summary_sheet(
        workbook,
        "寿险汇总-含外协（2026）",
        [
            [date(2026, 2, 1), 480, 606, 126, 387, 986, 5.2, 861716, 1266448, 118.5, 1, 1],
        ],
    )
    _append_current_external_sheet(
        workbook,
        "寿险对外数据（2026）",
        "寿险",
        [
            ["1月", 121.6, 305, 1061, 836, 2116264, 3181423, 0.000942699, 4.34, 1995, 1, 0.9, 0.7, 0.95, 10, 8, 0.8],
        ],
    )
    _append_current_team_sheet(
        workbook,
        "各团队数据（寿险）-2026",
        {
            1: [["A团队", 12, 60, 90, 45, 20, 6000, 9800, 6, 15, 1, 7.5, 1, 3.33, 2, 0.0033, 2, 1000, 2, 1633.33, 2]],
            2: [["A团队", 12, 70, 100, 50, 22, 6800, 11000, 6.5, 15.38, 1, 7.69, 1, 3.38, 2, 0.0032, 1, 1046.15, 2, 1692.31, 2]],
        },
    )

    content = io.BytesIO()
    workbook.save(content)
    workbook.close()
    return content.getvalue()


def _build_team_note_workbook_bytes() -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)

    _append_current_summary_sheet(
        workbook,
        "\u5bff\u9669\u6c47\u603b\u6570\u636e-\u542b\u5916\u534f\uff082026\uff09",
        [
            [1, 1061, 1153, 305, 836, 1995, 4.34, 2116264, 3181423, 121.6, 1, 1],
            [2, 480, 606, 126, 387, 986, 5.2, 861716, 1266448, 118.5, 1, 1],
        ],
    )
    _append_current_external_sheet(
        workbook,
        "\u5bff\u9669\u5bf9\u5916\u6570\u636e\uff082026\uff09",
        "\u5bff\u9669",
        [
            ["1\u6708", 121.6, 305, 1061, 836, 2116264, 3181423, 0.000942699, 4.34, 1995, 1, 0.9, 0.7, 0.95, 10, 8, 0.8],
            ["2\u6708", 118.5, 126, 480, 387, 861716, 1266448, 0.001144228, 5.2, 986, 0, 1, 0.72, 0.96, 10, 9, 0.9],
        ],
    )

    sheet = workbook.create_sheet("\u5404\u56e2\u961f\u6570\u636e\uff08\u5bff\u9669\uff09-2026")
    team_header = [
        "\u56e2\u961f",
        "\u7cfb\u7edf\u4e2a\u6570",
        "\u540c\u6b65\u4efb\u52a1\u6570",
        "\u540c\u6b65\u4efb\u52a1\u6570+\n\u56de\u5f52\u4efb\u52a1\u6570",
        "\u603b\u9700\u6c42\u6570",
        "\u603bBUG\u6570\n\uff08\u5168\u4efb\u52a1\uff09",
        "\u8bbe\u8ba1\u6848\u4f8b\u6570\n\uff08\u540c\u6b65\u4efb\u52a1\uff09",
        "\u6267\u884c\u6848\u4f8b\u6570\n\uff08\u540c\u6b65+\u56de\u5f52(\u624b\u5de5&\u81ea\u52a8\u5316)+\u5b89\u5168\uff09",
        "\u4eba\u6708\u6570",
        "\u4eba\u5747\u4efb\u52a1",
        None,
        "\u4eba\u5747\u9700\u6c42\u6570",
        None,
        "\u4eba\u5747\u7f3a\u9677\u6570",
        None,
        "\u7f3a\u9677\u7387",
        None,
        "\u5e73\u5747\u8bbe\u8ba1\u6848\u4f8b\u6570",
        None,
        "\u5e73\u5747\u6267\u884c\u6848\u4f8b\u6570",
        None,
    ]
    team_sub_header = [None, None, None, None, None, None, None, None, None, "\u6570\u91cf", "\u6392\u540d", "\u6570\u91cf", "\u6392\u540d", "\u6570\u91cf", "\u6392\u540d", "\u6570\u91cf", "\u6392\u540d", "\u6570\u91cf", "\u6392\u540d", "\u6570\u91cf", "\u6392\u540d"]
    month_one_row = ["A\u56e2\u961f", 12, 60, 90, 45, 20, 6000, 9800, 6, 15, 1, 7.5, 1, 3.33, 2, 0.0033, 2, 1000, 2, 1633.33, 2]
    month_two_row = ["A\u56e2\u961f", 12, 70, 100, 50, 22, 6800, 11000, 6.5, 15.38, 1, 7.69, 1, 3.38, 2, 0.0032, 1, 1046.15, 2, 1692.31, 2]

    sheet.append([None] * 21)
    sheet.append(["2026\u5e741\u6708\u4efd"] + [None] * 20)
    sheet.append(team_header)
    sheet.append(team_sub_header)
    sheet.append(month_one_row)
    sheet.append(["\u6ce8\u91ca\uff1a\u4e0b\u4e00\u4e2a\u6708\u4efd\u5757\u524d\u63d2\u5165\u5907\u6ce8\u884c"])
    sheet.append(["2026\u5e742\u6708\u4efd"] + [None] * 20)
    sheet.append(team_header)
    sheet.append(team_sub_header)
    sheet.append(month_two_row)

    content = io.BytesIO()
    workbook.save(content)
    workbook.close()
    return content.getvalue()


def test_analyze_performance_workbook_builds_history_and_current_sections():
    result = analyze_performance_workbook(_build_workbook_bytes())

    assert sorted(result["available_businesses"]) == ["健康险", "寿险"]
    assert result["current_year"] == 2026

    life = result["businesses"]["寿险"]
    assert life["history"]["latest_year"] == 2025
    assert [row["year"] for row in life["history"]["yearly_summary"]] == [2023, 2024, 2025]
    assert life["history"]["yearly_summary"][2]["defect_count"] == pytest.approx(3041)
    assert life["history"]["yearly_summary"][0]["defect_rate"] == pytest.approx(1380.17 / 807913.5, rel=1e-3)
    assert life["history"]["yearly_summary_table"]["headers"][3] == "发布总次数"
    assert life["history"]["yearly_summary_table"]["headers"][5] == "缺陷数 除性能、代码扫描外所有任务类型（SIT+FT）"
    assert "" not in life["history"]["yearly_summary_table"]["headers"]
    assert life["history"]["yearly_summary_table"]["rows"][0][0] == "2023年平均数据"
    assert life["history"]["yearly_summary_table"]["rows"][0][5] == 1380.17
    assert life["history"]["yearly_summary_table"]["rows"][0][6] == 4.75
    assert life["history"]["yearly_per_capita_table"]["headers"][4] == "人均总缺陷数"
    assert "" not in life["history"]["yearly_per_capita_table"]["headers"]
    assert life["history"]["yearly_per_capita_table"]["rows"][0][0] == "23年人均任务数"
    assert life["history"]["yearly_per_capita_table"]["rows"][0][5] == pytest.approx(0.00171)

    assert life["current"]["year"] == 2026
    assert life["current"]["latest_month"] == {"year": 2026, "month": 2, "month_label": "2月"}
    assert len(life["current"]["month_options"]) == 12
    assert life["current"]["month_options"][0]["disabled"] is False
    assert life["current"]["month_options"][1]["disabled"] is False
    assert life["current"]["month_options"][2]["disabled"] is True
    assert life["current"]["months"][1]["summary"]["sync_tasks"] == 1061
    assert life["current"]["months"][2]["external"]["automation_coverage"] == pytest.approx(0.72)
    assert life["current"]["months"][2]["team_snapshot"]["teams"][0]["team_name"] == "A团队"


def test_analyze_performance_workbook_ignores_non_whitelist_and_uses_latest_current_year():
    result = analyze_performance_workbook(_build_workbook_bytes())

    assert "寿险对外数据（2025）" in result["sheet_names"]
    assert "指标口径定义" in result["sheet_names"]
    assert result["businesses"]["寿险"]["current"]["year"] == 2026
    assert result["businesses"]["寿险"]["current"]["months"][1]["summary"]["year"] == 2026
    assert result["businesses"]["寿险"]["current"]["months"][1]["summary"]["sync_tasks"] != 9999
    assert result["businesses"]["健康险"]["history"]["available_years"] == [2024, 2025]


def test_analyze_performance_workbook_merges_current_months_from_multiple_summary_sheets():
    result = analyze_performance_workbook(_build_split_current_summary_workbook_bytes())

    life = result["businesses"]["寿险"]
    assert life["current"]["year"] == 2026
    assert life["current"]["latest_month"] == {"year": 2026, "month": 2, "month_label": "2月"}
    assert life["current"]["month_options"][0]["disabled"] is False
    assert life["current"]["month_options"][1]["disabled"] is False
    assert life["current"]["months"][2]["summary"]["sync_tasks"] == 480
    assert life["current"]["months"][2]["summary"]["demand_count"] == 387
    assert life["current"]["months"][2]["team_snapshot"]["month"] == 2


def test_analyze_performance_workbook_keeps_following_team_month_after_note_row():
    result = analyze_performance_workbook(_build_team_note_workbook_bytes())

    life = result["businesses"]["\u5bff\u9669"]
    assert life["current"]["latest_month"] == {"year": 2026, "month": 2, "month_label": "\u0032\u6708"}
    assert life["current"]["month_options"][1]["disabled"] is False
    assert life["current"]["months"][2]["summary"]["sync_tasks"] == 480
    assert life["current"]["months"][2]["external"]["defect_count"] == 986
    assert life["current"]["months"][2]["team_snapshot"]["month"] == 2
    assert life["current"]["months"][2]["team_snapshot"]["teams"][0]["team_name"] == "A\u56e2\u961f"


def test_upload_and_analyze_performance_workbook(client):
    workbook_bytes = _build_workbook_bytes()

    upload_resp = client.post(
        "/api/performance-analysis-files",
        files={
            "file": (
                "efficiency-dashboard.xlsx",
                workbook_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert upload_resp.status_code == 200
    upload_data = upload_resp.json()
    assert upload_data["success"] is True
    assert upload_data["data"]["sheet_count"] == 11

    file_id = upload_data["data"]["id"]
    analysis_resp = client.get(f"/api/performance-analysis-files/{file_id}/analysis")
    assert analysis_resp.status_code == 200

    analysis_data = analysis_resp.json()
    assert analysis_data["success"] is True
    assert analysis_data["data"]["current_year"] == 2026
    assert analysis_data["data"]["businesses"]["寿险"]["history"]["latest_year"] == 2025
    assert analysis_data["data"]["businesses"]["寿险"]["current"]["latest_month"]["month"] == 2
    assert len(analysis_data["data"]["businesses"]["寿险"]["current"]["month_options"]) == 12
    assert analysis_data["data"]["businesses"]["寿险"]["current"]["months"]["2"]["team_snapshot"]["month"] == 2
